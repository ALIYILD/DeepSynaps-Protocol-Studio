from __future__ import annotations
"""Export the full evidence DB to a doctor-readable Excel file.

Sheet 1 Papers    — one row per paper with PMID + DOI + OA links.
Sheet 2 Trials    — NCT + interventions (the real protocol parameter source).
Sheet 3 Devices   — FDA PMA/510(k)/HDE rows.
Sheet 4 Summary   — counts per indication.
Usage:
    python3 services/evidence-pipeline/scripts/export_matrix_xlsx.py [OUT]
"""
import json
import sys
from pathlib import Path

HERE = Path(__file__).resolve()
PIPELINE = HERE.parents[1]
sys.path.insert(0, str(PIPELINE))
import db
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


GRADE_FILL = {
    "A": PatternFill("solid", fgColor="C6EFCE"),
    "B": PatternFill("solid", fgColor="E2EFDA"),
    "C": PatternFill("solid", fgColor="FFF2CC"),
    "D": PatternFill("solid", fgColor="FCE4D6"),
    "E": PatternFill("solid", fgColor="F8CBAD"),
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, widths):
    for i, c in enumerate(ws[1], 1):
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for i, w in enumerate(widths, 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = w


def build(out_path: str) -> None:
    conn = db.connect()
    wb = Workbook()

    # ── Papers ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Papers"
    headers = [
        "Title", "First author", "Year", "Journal", "Study type",
        "Cites", "OA?", "Grade", "Modality", "Condition", "Indication",
        "PMID (link)", "DOI (link)", "OA PDF (link)",
    ]
    ws.append(headers)

    sql = (
        "SELECT p.title, p.authors_json, p.year, p.journal, p.pub_types_json, "
        "p.cited_by_count, p.is_oa, p.oa_url, p.pmid, p.doi, "
        "i.evidence_grade, i.modality, i.condition, i.slug "
        "FROM papers p "
        "JOIN paper_indications pi ON pi.paper_id = p.id "
        "JOIN indications i ON i.id = pi.indication_id "
        "ORDER BY p.year DESC, p.cited_by_count DESC"
    )
    for r in conn.execute(sql).fetchall():
        authors = json.loads(r["authors_json"] or "[]")
        first_author = authors[0] if authors else ""
        pub_types = json.loads(r["pub_types_json"] or "[]")
        study = ""
        for t in ("Meta-Analysis", "Systematic Review", "Practice Guideline",
                  "Randomized Controlled Trial", "Clinical Trial", "Review", "Case Reports"):
            if t in pub_types:
                study = t; break
        ws.append([
            r["title"],
            first_author + (" et al." if len(authors) > 1 else ""),
            r["year"],
            r["journal"],
            study,
            r["cited_by_count"] or 0,
            "Y" if r["is_oa"] else "N",
            r["evidence_grade"] or "",
            r["modality"],
            r["condition"],
            r["slug"],
            f'https://pubmed.ncbi.nlm.nih.gov/{r["pmid"]}' if r["pmid"] else "",
            f'https://doi.org/{r["doi"]}' if r["doi"] else "",
            r["oa_url"] or "",
        ])
        row_idx = ws.max_row
        if r["evidence_grade"] in GRADE_FILL:
            ws.cell(row=row_idx, column=8).fill = GRADE_FILL[r["evidence_grade"]]
        for col in (12, 13, 14):  # hyperlinks
            cell = ws.cell(row=row_idx, column=col)
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _style_header(ws, [42, 20, 8, 26, 22, 7, 6, 7, 12, 26, 22, 32, 32, 36])

    # ── Trials ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Trials")
    ws2.append([
        "NCT ID (link)", "Title", "Phase", "Status", "Enrollment",
        "Sponsor", "Conditions", "Intervention summary", "Modality", "Indication", "Last update",
    ])
    sql2 = (
        "SELECT t.nct_id, t.title, t.phase, t.status, t.enrollment, t.sponsor, "
        "t.conditions_json, t.interventions_json, t.last_update, i.modality, i.slug "
        "FROM trials t "
        "JOIN trial_indications ti ON ti.trial_id = t.id "
        "JOIN indications i ON i.id = ti.indication_id "
        "ORDER BY t.last_update DESC"
    )
    for r in conn.execute(sql2).fetchall():
        conds = json.loads(r["conditions_json"] or "[]")
        ivs = json.loads(r["interventions_json"] or "[]")
        iv_summary = "; ".join(
            (iv.get("name") or "") + (f" ({iv.get('type')})" if iv.get("type") else "")
            for iv in ivs[:3]
        )
        ws2.append([
            r["nct_id"], r["title"], r["phase"], r["status"], r["enrollment"],
            r["sponsor"], ", ".join(conds), iv_summary,
            r["modality"], r["slug"], r["last_update"],
        ])
        if r["nct_id"]:
            cell = ws2.cell(row=ws2.max_row, column=1)
            cell.hyperlink = f"https://clinicaltrials.gov/study/{r['nct_id']}"
            cell.font = Font(color="0563C1", underline="single")
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    _style_header(ws2, [14, 48, 10, 18, 10, 28, 28, 46, 12, 22, 14])

    # ── Devices ───────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("FDA Devices")
    ws3.append([
        "Kind", "Number", "Product code", "Applicant", "Trade name",
        "Decision date", "Modality", "Indication",
    ])
    sql3 = (
        "SELECT d.kind, d.number, d.product_code, d.applicant, d.trade_name, d.decision_date, "
        "i.modality, i.slug "
        "FROM devices d "
        "JOIN device_indications di ON di.device_id = d.id "
        "JOIN indications i ON i.id = di.indication_id "
        "ORDER BY d.decision_date DESC"
    )
    for r in conn.execute(sql3).fetchall():
        ws3.append([
            r["kind"].upper(), r["number"], r["product_code"], r["applicant"],
            r["trade_name"], r["decision_date"], r["modality"], r["slug"],
        ])
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = ws3.dimensions
    _style_header(ws3, [8, 12, 12, 28, 48, 14, 12, 22])

    # ── Summary ───────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Summary")
    ws4.append(["Indication", "Modality", "Grade", "Papers", "Trials", "Devices"])
    sql4 = (
        "SELECT i.slug, i.modality, i.evidence_grade, "
        "(SELECT count(*) FROM paper_indications pi WHERE pi.indication_id=i.id) AS papers, "
        "(SELECT count(*) FROM trial_indications ti WHERE ti.indication_id=i.id) AS trials, "
        "(SELECT count(*) FROM device_indications di WHERE di.indication_id=i.id) AS devices "
        "FROM indications i ORDER BY papers DESC"
    )
    for r in conn.execute(sql4).fetchall():
        ws4.append([r["slug"], r["modality"], r["evidence_grade"] or "", r["papers"], r["trials"], r["devices"]])
        if r["evidence_grade"] in GRADE_FILL:
            ws4.cell(row=ws4.max_row, column=3).fill = GRADE_FILL[r["evidence_grade"]]
    _style_header(ws4, [30, 14, 8, 10, 10, 10])

    wb.save(out_path)
    print(f"saved {out_path}")
    print(f"papers sheet: {ws.max_row - 1} rows")
    print(f"trials sheet: {ws2.max_row - 1} rows")
    print(f"devices sheet: {ws3.max_row - 1} rows")


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else str(Path.home() / "Desktop" / "deepsynaps-evidence.xlsx")
    build(out)
