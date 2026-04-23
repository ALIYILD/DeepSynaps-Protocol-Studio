from __future__ import annotations

import argparse
import csv
import json
import re
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path
from statistics import mean

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from indications_seed import MODALITY_PRODUCT_CODES, SEED
from sources import ctgov, openfda


HEADER_FILL = PatternFill("solid", fgColor="16324F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D5DCE5")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
EUROPEPMC_SEARCH_URL = "https://www.ebi.ac.uk/europepmc/webservices/rest/search"
PUBMED_ESUMMARY_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi"
CROSSREF_WORKS_URL = "https://api.crossref.org/works/"


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, fieldnames: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def split_tags(value: str | None) -> list[str]:
    if not value:
        return []
    normalized = value.replace("|", ";")
    return [part.strip() for part in normalized.split(";") if part.strip()]


def safe_int(value: str | None) -> int:
    if not value:
        return 0
    try:
        return int(float(value))
    except Exception:
        return 0


def normalize_text(value: str | None, limit: int | None = None) -> str:
    text = " ".join((value or "").split())
    if limit and len(text) > limit:
        return text[: limit - 1].rstrip() + "…"
    return text


def cleanup_author_token(token: str) -> str:
    token = normalize_text(token.strip(" ,.;"))
    if not token:
        return ""
    if re.fullmatch(r"[A-Z][A-Z \-']{3,}", token):
        token = token.title()
    return re.sub(r"\s+", " ", token)


def normalize_authors(raw: str | None) -> tuple[str, str, str, int]:
    raw = normalize_text(raw)
    if not raw:
        return "", "missing", "", 0
    if raw.strip(" ,.;") == "":
        return "", "placeholder_only", "", 0
    tokens: list[str] = []
    for splitter in [r";", r","]:
        candidate = []
        for part in re.split(splitter, raw):
            cleaned = cleanup_author_token(part)
            if cleaned:
                candidate.append(cleaned)
        if candidate:
            tokens = candidate
            break
    deduped = []
    for token in tokens:
        if token not in deduped:
            deduped.append(token)
    if not deduped:
        return "", "unparseable", "", 0
    normalized = "; ".join(deduped)
    return normalized, "clean", deduped[0], len(deduped)


def separate_group_authors(raw: str | None) -> tuple[str, str, str]:
    raw = normalize_text(raw)
    if not raw:
        return "", "", "missing"
    parts = [cleanup_author_token(part) for part in re.split(r";", raw)]
    if len(parts) == 1:
        parts = [cleanup_author_token(part) for part in raw.split(",")]
    parts = [part for part in parts if part]
    person_parts: list[str] = []
    group_parts: list[str] = []
    for part in parts:
        lowered = part.lower()
        if any(
            keyword in lowered
            for keyword in [
                "group",
                "consortium",
                "consensus",
                "committee",
                "working group",
                "collaboration",
                "society",
                "network",
                "task force",
                "investigators",
                "panel",
                "association",
                "initiative",
                "team",
                "on behalf of",
                "members of the",
            ]
        ):
            group_parts.append(part)
        else:
            person_parts.append(part)
    notes = "separated"
    if group_parts and any("proceedings of the meeting" in part.lower() for part in group_parts):
        notes = "separated_with_meeting_metadata"
    elif not group_parts:
        notes = "no_group_detected"
    return "; ".join(person_parts), "; ".join(group_parts), notes


def normalize_journal_name(raw: str | None, source: str | None = None) -> tuple[str, str]:
    journal = normalize_text(raw)
    if not journal:
        return "", "missing"
    journal = re.sub(r"\s*[\.,;:]+$", "", journal)
    if journal.isupper() and len(journal) > 6:
        journal = journal.title()
    quality = "provided_pmc" if source == "PMC" else "provided"
    return journal, quality


def europepmc_query_for_row(row: dict[str, str]) -> str | None:
    source = normalize_text(row.get("source"))
    source_id = normalize_text(row.get("id"))
    if source and source_id:
        return f'EXT_ID:"{source_id}" AND SRC:"{source}"'
    if normalize_text(row.get("pmcid")):
        return f'PMCID:"{normalize_text(row.get("pmcid"))}"'
    if normalize_text(row.get("pmid")):
        return f'EXT_ID:"{normalize_text(row.get("pmid"))}" AND SRC:"MED"'
    if normalize_text(row.get("doi")):
        return f'DOI:"{normalize_text(row.get("doi"))}"'
    return None


def backfill_journal_for_row(row: dict[str, str]) -> tuple[str, str]:
    query = europepmc_query_for_row(row)
    if not query:
        return "", "missing_identifier"
    params = urllib.parse.urlencode({"query": query, "format": "json", "pageSize": "1", "resultType": "core"})
    url = f"{EUROPEPMC_SEARCH_URL}?{params}"
    try:
        with urllib.request.urlopen(url, timeout=30) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except Exception:
        return "", "fetch_error"
    results = payload.get("resultList", {}).get("result", [])
    if not results:
        return "", "not_found"
    record = results[0]
    journal = (
        record.get("journalTitle")
        or record.get("journal")
        or record.get("journalInfo", {}).get("journal", {}).get("title")
        or ""
    )
    return normalize_text(journal), "europepmc_backfill" if journal else "not_found"


def fetch_json(url: str, *, timeout: int = 30) -> dict:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "DeepSynaps-Protocol-Studio/1.0 (journal backfill)",
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def backfill_journal_pubmed(row: dict[str, str]) -> tuple[str, str]:
    pmid = normalize_text(row.get("pmid"))
    if not pmid:
        return "", "missing_pmid"
    params = urllib.parse.urlencode({"db": "pubmed", "id": pmid, "retmode": "json"})
    try:
        payload = fetch_json(f"{PUBMED_ESUMMARY_URL}?{params}")
    except Exception:
        return "", "pubmed_fetch_error"
    result = (payload.get("result") or {}).get(pmid) or {}
    journal = (
        result.get("fulljournalname")
        or result.get("source")
        or result.get("journalname")
        or ""
    )
    return normalize_text(journal), "pubmed_esummary" if journal else "pubmed_not_found"


def backfill_journal_crossref(row: dict[str, str]) -> tuple[str, str]:
    doi = normalize_text(row.get("doi"))
    if not doi:
        return "", "missing_doi"
    doi_path = urllib.parse.quote(doi, safe="")
    try:
        payload = fetch_json(f"{CROSSREF_WORKS_URL}{doi_path}")
    except Exception:
        return "", "crossref_fetch_error"
    message = payload.get("message") or {}
    container_titles = message.get("container-title") or []
    journal = container_titles[0] if container_titles else ""
    return normalize_text(journal), "crossref" if journal else "crossref_not_found"


def attempt_journal_backfill(row: dict[str, str]) -> tuple[str, str]:
    for fetcher in (backfill_journal_pubmed, backfill_journal_crossref, backfill_journal_for_row):
        journal, source = fetcher(row)
        if journal:
            return journal, source
    if normalize_text(row.get("pmid")):
        return "", "pubmed_crossref_europepmc_not_found"
    if normalize_text(row.get("doi")):
        return "", "crossref_europepmc_not_found"
    return "", "missing_identifier"


def apply_bounded_journal_backfill(
    enriched_rows: list[dict],
    derived_dir: Path,
    *,
    limit: int,
    min_priority_score: int,
    min_confidence_score: int,
) -> dict[str, int]:
    attempt_rows: list[dict[str, str]] = []
    candidates = [
        row
        for row in enriched_rows
        if not normalize_text(row.get("journal_normalized"))
        and (
            normalize_text(row.get("pmid"))
            or normalize_text(row.get("doi"))
            or normalize_text(row.get("pmcid"))
            or (normalize_text(row.get("source")) and normalize_text(row.get("id")))
        )
        and safe_int(str(row.get("priority_score"))) >= min_priority_score
        and safe_int(str(row.get("paper_confidence_score"))) >= min_confidence_score
    ]
    candidates.sort(
        key=lambda row: (
            safe_int(str(row.get("priority_score"))),
            safe_int(str(row.get("paper_confidence_score"))),
            safe_int(str(row.get("cited_by_count"))),
        ),
        reverse=True,
    )

    attempted = 0
    updated = 0
    for row in candidates[:limit]:
        attempted += 1
        journal, source = attempt_journal_backfill(row)
        if journal:
            row["journal"] = journal
            row["journal_normalized"], row["journal_quality_flag"] = normalize_journal_name(journal, source)
            row["journal_backfill_source"] = source
            updated += 1
        else:
            row["journal_backfill_source"] = source
        attempt_rows.append(
            {
                "paper_key": row.get("paper_key", ""),
                "title": row.get("title", ""),
                "pmid": row.get("pmid", ""),
                "doi": row.get("doi", ""),
                "priority_score": str(row.get("priority_score", "")),
                "paper_confidence_score_before": str(row.get("paper_confidence_score", "")),
                "journal_after": row.get("journal", ""),
                "journal_normalized_after": row.get("journal_normalized", ""),
                "journal_backfill_source": row.get("journal_backfill_source", ""),
            }
        )

    audit_path = derived_dir / "neuromodulation_journal_backfill.csv"
    if attempt_rows:
        write_csv(audit_path, list(attempt_rows[0].keys()), attempt_rows)
    return {
        "candidate_count": len(candidates),
        "attempted": attempted,
        "updated": updated,
    }


def modality_seed_metadata() -> dict[str, dict[str, set[str]]]:
    out: dict[str, dict[str, set[str]]] = defaultdict(
        lambda: {"trial_queries": set(), "applicants": set(), "conditions": set(), "product_codes": set()}
    )
    for entry in SEED:
        bucket = out[entry["modality"]]
        bucket["trial_queries"].add(entry["trial_q"])
        bucket["conditions"].add(entry["condition"])
        bucket["applicants"].update(entry.get("fda_applicants") or [])
        bucket["product_codes"].update(MODALITY_PRODUCT_CODES.get(entry["modality"]) or [])
    return out


AI_TO_SEED_MODALITY = {
    "deep_brain_stimulation": "DBS",
    "responsive_neurostimulation": "RNS",
    "vagus_nerve_stimulation": "VNS",
    "spinal_cord_stimulation": "SCS",
    "dorsal_root_ganglion_stimulation": "DRG",
    "sacral_neuromodulation": "SNM",
    "hypoglossal_nerve_stimulation": "HNS",
    "transcranial_magnetic_stimulation": "rTMS",
    "transcranial_direct_current_stimulation": "tDCS",
    "focused_ultrasound_neuromodulation": "MRgFUS",
}

CONDITION_QUERY_LABELS = {
    "chronic_pain": ["chronic pain", "neuropathic pain", "failed back surgery syndrome", "painful diabetic neuropathy"],
    "parkinsons_disease": ["Parkinson disease", "Parkinson's disease"],
    "depression": ["depression", "major depressive disorder", "treatment resistant depression"],
    "epilepsy_seizures": ["epilepsy", "focal epilepsy", "seizures"],
    "stroke_rehabilitation": ["stroke rehabilitation", "post-stroke", "stroke"],
    "urinary_bladder_bowel": ["overactive bladder", "urinary retention", "fecal incontinence", "bladder dysfunction"],
    "alzheimers_dementia": ["Alzheimer disease", "dementia", "cognitive impairment"],
    "addiction_substance_use": ["substance use disorder", "addiction", "alcohol use disorder", "opioid use disorder"],
    "dystonia": ["dystonia"],
    "spinal_cord_injury": ["spinal cord injury"],
    "essential_tremor": ["essential tremor"],
    "multiple_sclerosis": ["multiple sclerosis"],
    "ocd": ["obsessive compulsive disorder", "OCD"],
    "anxiety": ["anxiety", "generalized anxiety disorder"],
    "brain_injury": ["traumatic brain injury", "brain injury"],
    "tinnitus": ["tinnitus"],
    "sleep_apnea": ["obstructive sleep apnea", "sleep apnea"],
    "ptsd": ["PTSD", "post-traumatic stress disorder"],
}

MODALITY_QUERY_LABELS = {
    "deep_brain_stimulation": ["deep brain stimulation", "DBS"],
    "responsive_neurostimulation": ["responsive neurostimulation", "RNS"],
    "vagus_nerve_stimulation": ["vagus nerve stimulation", "VNS"],
    "spinal_cord_stimulation": ["spinal cord stimulation", "SCS", "dorsal column stimulation"],
    "dorsal_root_ganglion_stimulation": ["dorsal root ganglion stimulation", "DRG stimulation"],
    "sacral_neuromodulation": ["sacral neuromodulation", "sacral nerve stimulation", "InterStim", "Axonics"],
    "hypoglossal_nerve_stimulation": ["hypoglossal nerve stimulation", "Inspire"],
    "transcranial_magnetic_stimulation": ["transcranial magnetic stimulation", "rTMS", "TMS", "iTBS"],
    "transcranial_direct_current_stimulation": ["transcranial direct current stimulation", "tDCS"],
    "focused_ultrasound_neuromodulation": ["focused ultrasound", "MRgFUS", "Exablate"],
    "peripheral_nerve_stimulation": ["peripheral nerve stimulation", "PNS"],
    "occipital_nerve_stimulation": ["occipital nerve stimulation"],
    "tibial_nerve_stimulation": ["tibial nerve stimulation", "PTNS", "TTNS"],
}

FDA_DEVICE_KEYWORDS = {
    "deep_brain_stimulation": ["deep brain", "dbs", "neurostimulator"],
    "responsive_neurostimulation": ["responsive neurostimulation", "rns"],
    "vagus_nerve_stimulation": ["vagus", "vns"],
    "spinal_cord_stimulation": ["spinal cord", "scs", "dorsal column"],
    "dorsal_root_ganglion_stimulation": ["dorsal root ganglion", "drg"],
    "sacral_neuromodulation": ["sacral", "interstim", "axonics"],
    "hypoglossal_nerve_stimulation": ["hypoglossal", "inspire", "sleep apnea"],
    "transcranial_magnetic_stimulation": ["transcranial magnetic", "tms", "rtms", "deep tms"],
    "transcranial_direct_current_stimulation": ["transcranial direct current", "tdcs"],
    "focused_ultrasound_neuromodulation": ["focused ultrasound", "mrgfus", "exablate"],
    "peripheral_nerve_stimulation": ["peripheral nerve stimulation", "pns"],
}

IRRELEVANT_TRIAL_TERMS = {
    "drug",
    "pharmacokinetic",
    "pharmacodynamics",
    "capsule",
    "tablet",
    "xanomeline",
    "trospium",
    "placebo",
}

TRIAL_STATUS_PRIORITY = {
    "RECRUITING": 5,
    "NOT_YET_RECRUITING": 4,
    "ACTIVE_NOT_RECRUITING": 4,
    "ENROLLING_BY_INVITATION": 3,
    "COMPLETED": 3,
    "SUSPENDED": 1,
    "WITHDRAWN": 0,
    "TERMINATED": 0,
    "UNKNOWN": 1,
}

FDA_PRODUCT_CODE_LABELS = {
    "GZB": "deep brain stimulation",
    "LGW": "responsive neurostimulation",
    "NHH": "vagus nerve stimulation",
    "LGX": "spinal cord stimulation",
    "QLX": "dorsal root ganglion stimulation",
    "MWW": "sacral neuromodulation",
    "MNQ": "hypoglossal nerve stimulation",
    "OBP": "transcranial magnetic stimulation",
    "QIH": "focused ultrasound neuromodulation",
}


def slug_to_phrase(value: str | None) -> str:
    return (value or "").replace("_", " ").strip()


def text_tokens(value: str | None) -> set[str]:
    return {token for token in re.findall(r"[a-z0-9]+", (value or "").lower()) if len(token) > 2}


def collect_text_fragments(payload: object) -> list[str]:
    if isinstance(payload, dict):
        out: list[str] = []
        for value in payload.values():
            out.extend(collect_text_fragments(value))
        return out
    if isinstance(payload, list):
        out = []
        for value in payload:
            out.extend(collect_text_fragments(value))
        return out
    if isinstance(payload, str):
        return [payload]
    return []


def parse_json_list(value: str | None) -> list:
    if not value:
        return []
    try:
        parsed = json.loads(value)
    except Exception:
        return []
    return parsed if isinstance(parsed, list) else []


def modality_terms(modality: str) -> list[str]:
    labels = MODALITY_QUERY_LABELS.get(modality, [])
    keywords = FDA_DEVICE_KEYWORDS.get(modality, [])
    return [term.lower() for term in (labels + keywords) if term]


def condition_terms(condition_slug: str) -> list[str]:
    labels = CONDITION_QUERY_LABELS.get(condition_slug, [])
    fallback = [slug_to_phrase(condition_slug)] if condition_slug else []
    return [term.lower() for term in (labels or fallback) if term]


def summarize_trial_row(row: dict) -> str:
    parts = [row.get("nct_id") or ""]
    title = normalize_text(row.get("title"), 90)
    if title:
        parts.append(title)
    status = row.get("overall_status") or ""
    phase = row.get("phase") or ""
    if status:
        parts.append(status)
    if phase:
        parts.append(phase)
    enrollment = row.get("enrollment") or ""
    if enrollment:
        parts.append(f"n={enrollment}")
    return " | ".join(part for part in parts if part)


def summarize_fda_row(row: dict) -> str:
    parts = [row.get("clearance_number") or ""]
    device_name = normalize_text(row.get("device_name"), 80)
    if device_name:
        parts.append(device_name)
    decision_date = row.get("decision_date") or ""
    if decision_date:
        parts.append(decision_date)
    product_code = row.get("product_code") or ""
    if product_code:
        parts.append(product_code)
    return " | ".join(part for part in parts if part)


def extract_protocol_parameters(text: str) -> dict[str, str]:
    lowered = (text or "").lower()
    frequency = re.findall(r"(\d+(?:\.\d+)?)\s*hz\b", lowered)
    pulse_width = re.findall(r"(\d+(?:\.\d+)?)\s*(?:μs|µs|us)\b", lowered)
    current = re.findall(r"(\d+(?:\.\d+)?)\s*ma\b", lowered)
    voltage = re.findall(r"(\d+(?:\.\d+)?)\s*v\b", lowered)
    duration = re.findall(r"(\d+(?:\.\d+)?)\s*(minutes?|mins?|hours?|days?|weeks?)\b", lowered)
    sessions = re.findall(r"(\d+)\s+sessions?\b", lowered)
    pulses = re.findall(r"(\d+)\s+pulses?\b", lowered)
    duty = re.findall(r"(\d+\s*s\s*on\s*/\s*\d+\s*s\s*off)", lowered)
    target_patterns = [
        "left ear",
        "right ear",
        "tragus",
        "cymba conchae",
        "left dlpfc",
        "right dlpfc",
        "dlpfc",
        "motor cortex",
        "m1",
        "subthalamic nucleus",
        "stn",
        "globus pallidus",
        "spinal cord",
        "dorsal root ganglion",
        "occipital nerve",
        "vagus nerve",
        "hypoglossal nerve",
        "tibial nerve",
        "sacral nerve",
    ]
    targets = [pattern for pattern in target_patterns if pattern in lowered]
    sham_control = "Y" if any(term in lowered for term in ["sham", "placebo", "control", "mock"]) else "N"
    return {
        "protocol_frequency_hz": "; ".join(dict.fromkeys(frequency)),
        "protocol_pulse_width_us": "; ".join(dict.fromkeys(pulse_width)),
        "protocol_current_ma": "; ".join(dict.fromkeys(current)),
        "protocol_voltage_v": "; ".join(dict.fromkeys(voltage)),
        "protocol_duration": "; ".join(f"{value} {unit}" for value, unit in duration[:8]),
        "protocol_session_count": "; ".join(dict.fromkeys(sessions)),
        "protocol_pulse_count": "; ".join(dict.fromkeys(pulses)),
        "protocol_duty_cycle": "; ".join(dict.fromkeys(duty)),
        "protocol_target_site": "; ".join(dict.fromkeys(targets)),
        "protocol_has_sham_control": sham_control,
    }


def summarize_protocol_parameters(row: dict) -> str:
    parts = []
    for label, key in [
        ("freq_hz", "protocol_frequency_hz"),
        ("pulse_width_us", "protocol_pulse_width_us"),
        ("current_ma", "protocol_current_ma"),
        ("duration", "protocol_duration"),
        ("sessions", "protocol_session_count"),
        ("duty", "protocol_duty_cycle"),
        ("target", "protocol_target_site"),
    ]:
        value = row.get(key) or ""
        if value:
            parts.append(f"{label}={value}")
    if (row.get("protocol_has_sham_control") or "") == "Y":
        parts.append("sham/control")
    return " | ".join(parts)


def score_trial_record(row: dict) -> tuple[int, list[str]]:
    modality = row.get("matched_modality") or ""
    conditions = parse_json_list(row.get("conditions_json"))
    interventions = parse_json_list(row.get("interventions_json"))
    text_parts = [
        row.get("title") or "",
        row.get("brief_summary") or "",
        " ".join(collect_text_fragments(conditions)),
        " ".join(collect_text_fragments(interventions)),
        row.get("query_used") or "",
    ]
    haystack = " ".join(text_parts).lower()
    score = 0
    reasons: list[str] = []

    for term in modality_terms(modality):
        if term and term in haystack:
            score += 4
            reasons.append(f"modality:{term}")
            break
    if any("device" == (item.get("type") or "").lower() for item in interventions if isinstance(item, dict)):
        score += 2
        reasons.append("device_intervention")
    if any("stimulation" in fragment.lower() or "neuromod" in fragment.lower() for fragment in collect_text_fragments(interventions)):
        score += 2
        reasons.append("stimulation_intervention")
    if any(term in haystack for term in ("sham stimulation", "deep tms", "taVNS".lower(), "tms", "dbs", "vns", "rns", "tdcs", "focused ultrasound")):
        score += 1
        reasons.append("named_modulation_signal")

    matched_conditions = 0
    for condition_slug in split_tags(row.get("matched_conditions")):
        for term in condition_terms(condition_slug):
            if term and term in haystack:
                matched_conditions += 1
                reasons.append(f"condition:{condition_slug}")
                break
    score += min(matched_conditions * 2, 6)

    if "healthy volunteers" in haystack:
        score -= 2
        reasons.append("healthy_volunteers")
    if all((item.get("type") or "").upper() == "DRUG" for item in interventions if isinstance(item, dict)) and interventions:
        score -= 4
        reasons.append("drug_only")
    if any(term in haystack for term in IRRELEVANT_TRIAL_TERMS):
        score -= 2
        reasons.append("drug_language")

    score += TRIAL_STATUS_PRIORITY.get((row.get("overall_status") or "").upper(), 0)
    enrollment = safe_int(row.get("enrollment"))
    if enrollment >= 100:
        score += 2
        reasons.append("enrollment_100_plus")
    elif enrollment >= 20:
        score += 1
        reasons.append("enrollment_20_plus")
    return score, reasons


def score_fda_record(row: dict) -> tuple[int, list[str]]:
    modality = row.get("matched_modality") or ""
    haystack = " ".join(
        [
            row.get("device_name") or "",
            row.get("generic_name") or "",
            row.get("applicant") or "",
            row.get("product_code") or "",
            row.get("query_used") or "",
        ]
    ).lower()
    score = 0
    reasons: list[str] = []

    for term in modality_terms(modality):
        if term and term in haystack:
            score += 4
            reasons.append(f"modality:{term}")
            break
    product_code = (row.get("product_code") or "").upper()
    if product_code and product_code in (MODALITY_PRODUCT_CODES.get(AI_TO_SEED_MODALITY.get(modality, ""), set()) or set()):
        score += 4
        reasons.append(f"product_code:{product_code}")
    elif product_code in FDA_PRODUCT_CODE_LABELS:
        score += 1
        reasons.append(f"known_code:{product_code}")
    if filter_fda_record_for_modality(row, modality):
        score += 2
        reasons.append("keyword_match")
    if row.get("advisory_committee", "").lower() == "neurology":
        score += 1
        reasons.append("neurology_committee")
    return score, reasons


def refine_trial_rows(trial_rows: list[dict]) -> list[dict]:
    refined: list[dict] = []
    for row in trial_rows:
        row = dict(row)
        protocol_text = " ".join(
            [
                row.get("title") or "",
                row.get("brief_summary") or "",
                row.get("interventions_json") or "",
                row.get("outcomes_json") or "",
            ]
        )
        row.update(extract_protocol_parameters(protocol_text))
        row["protocol_parameter_summary"] = summarize_protocol_parameters(row)
        score, reasons = score_trial_record(row)
        row["relevance_score"] = score
        row["relevance_bucket"] = "high" if score >= 10 else "medium" if score >= 6 else "low"
        row["relevance_reasons"] = "; ".join(reasons[:8])
        if score >= 6:
            refined.append(row)
    refined.sort(
        key=lambda item: (
            safe_int(str(item.get("relevance_score"))),
            TRIAL_STATUS_PRIORITY.get((item.get("overall_status") or "").upper(), 0),
            safe_int(str(item.get("enrollment"))),
            item.get("last_update") or "",
            item.get("nct_id") or "",
        ),
        reverse=True,
    )
    return refined


def refine_fda_rows(fda_rows: list[dict]) -> list[dict]:
    refined: list[dict] = []
    for row in fda_rows:
        row = dict(row)
        score, reasons = score_fda_record(row)
        row["relevance_score"] = score
        row["relevance_bucket"] = "high" if score >= 8 else "medium" if score >= 5 else "low"
        row["relevance_reasons"] = "; ".join(reasons[:8])
        if score >= 5:
            refined.append(row)
    refined.sort(
        key=lambda item: (
            safe_int(str(item.get("relevance_score"))),
            item.get("decision_date") or "",
            item.get("clearance_number") or "",
        ),
        reverse=True,
    )
    return refined


def attach_paper_level_external_signals(enriched_rows: list[dict], trial_rows: list[dict], fda_rows: list[dict]) -> None:
    trials_by_modality: dict[str, list[dict]] = defaultdict(list)
    fda_by_modality: dict[str, list[dict]] = defaultdict(list)
    paper_signal_cache: dict[tuple, dict[str, str | int]] = {}
    for row in trial_rows:
        trials_by_modality[row.get("matched_modality") or ""].append(row)
    for row in fda_rows:
        fda_by_modality[row.get("matched_modality") or ""].append(row)

    for row in enriched_rows:
        modality = row.get("primary_modality") or ""
        if not modality:
            row["trial_match_count"] = 0
            row["trial_signal_score"] = 0
            row["trial_top_nct_ids"] = ""
            row["trial_summary"] = ""
            row["trial_protocol_parameter_summary"] = ""
            row["fda_match_count"] = 0
            row["fda_signal_score"] = 0
            row["fda_top_clearances"] = ""
            row["fda_summary"] = ""
            row["regulatory_clinical_signal"] = ""
            continue

        indication_slugs = split_tags(row.get("indication_tags"))
        paper_summary = (row.get("research_summary") or "").lower()
        paper_targets = set(split_tags(row.get("target_tags")))
        paper_parameters = set(split_tags(row.get("parameter_signal_tags")))
        summary_flags = tuple(sorted(term for term in ["randomized", "sham", "double-blind"] if term in paper_summary))
        cache_key = (
            modality,
            tuple(sorted(indication_slugs)),
            tuple(sorted(paper_targets)),
            tuple(sorted(paper_parameters)),
            summary_flags,
        )
        cached = paper_signal_cache.get(cache_key)
        if cached is not None:
            row.update({k: str(v) if isinstance(v, int) else v for k, v in cached.items()})
            continue

        trial_candidates: list[tuple[int, dict]] = []
        for trial in trials_by_modality.get(modality, []):
            candidate_score = safe_int(str(trial.get("relevance_score")))
            haystack = " ".join(
                [
                    trial.get("title") or "",
                    trial.get("brief_summary") or "",
                    trial.get("conditions_json") or "",
                    trial.get("interventions_json") or "",
                ]
            ).lower()
            if summary_flags and any(term in haystack for term in summary_flags):
                candidate_score += 1
            for condition_slug in indication_slugs:
                if any(term in haystack for term in condition_terms(condition_slug)):
                    candidate_score += 4
            trial_target = (trial.get("protocol_target_site") or "").lower()
            if paper_targets and any(target in trial_target for target in paper_targets):
                candidate_score += 2
            protocol_summary = (trial.get("protocol_parameter_summary") or "").lower()
            if paper_parameters and any(parameter in protocol_summary for parameter in paper_parameters):
                candidate_score += 1
            trial_candidates.append((candidate_score, trial))
        trial_candidates.sort(
            key=lambda item: (
                item[0],
                TRIAL_STATUS_PRIORITY.get((item[1].get("overall_status") or "").upper(), 0),
                safe_int(str(item[1].get("enrollment"))),
                item[1].get("last_update") or "",
            ),
            reverse=True,
        )
        selected_trials = [(score, trial) for score, trial in trial_candidates if score >= 8][:3]
        top_trials = [item[1] for item in selected_trials]

        fda_candidates: list[tuple[int, dict]] = []
        for device in fda_by_modality.get(modality, []):
            candidate_score = safe_int(str(device.get("relevance_score")))
            product_code = (device.get("product_code") or "").upper()
            if product_code and product_code in (MODALITY_PRODUCT_CODES.get(AI_TO_SEED_MODALITY.get(modality, ""), set()) or set()):
                candidate_score += 2
            device_haystack = " ".join(
                [
                    device.get("device_name") or "",
                    device.get("generic_name") or "",
                    device.get("applicant") or "",
                ]
            ).lower()
            for condition_slug in indication_slugs:
                if any(term in device_haystack for term in condition_terms(condition_slug)):
                    candidate_score += 2
            fda_candidates.append((candidate_score, device))
        fda_candidates.sort(
            key=lambda item: (item[0], item[1].get("decision_date") or "", item[1].get("clearance_number") or ""),
            reverse=True,
        )
        selected_devices = [(score, device) for score, device in fda_candidates if score >= 7][:3]
        top_devices = [item[1] for item in selected_devices]

        signal_payload = {
            "trial_match_count": len(top_trials),
            "trial_signal_score": sum(score for score, _trial in selected_trials),
            "trial_top_nct_ids": "; ".join(trial.get("nct_id") or "" for trial in top_trials),
            "trial_summary": " || ".join(summarize_trial_row(trial) for trial in top_trials),
            "trial_protocol_parameter_summary": " || ".join(
                trial.get("protocol_parameter_summary") or "" for trial in top_trials if trial.get("protocol_parameter_summary")
            ),
            "fda_match_count": len(top_devices),
            "fda_signal_score": sum(score for score, _device in selected_devices),
            "fda_top_clearances": "; ".join(device.get("clearance_number") or "" for device in top_devices),
            "fda_summary": " || ".join(summarize_fda_row(device) for device in top_devices),
            "regulatory_clinical_signal": " | ".join(
            part
            for part in [
                    f"{len(top_trials)} prioritized trials / score {sum(score for score, _trial in selected_trials)}" if top_trials else "",
                    f"{len(top_devices)} relevant 510(k) devices / score {sum(score for score, _device in selected_devices)}" if top_devices else "",
            ]
            if part
            ),
        }
        paper_signal_cache[cache_key] = signal_payload
        row.update({k: str(v) if isinstance(v, int) else v for k, v in signal_payload.items()})


def top_modality_condition_pairs(bundle_root: Path, limit: int = 30) -> list[tuple[str, str, int]]:
    path = bundle_root / "neuromodulation_master_database_enriched.csv"
    counts: Counter[tuple[str, str]] = Counter()
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            modality = (row.get("primary_modality") or "").strip()
            if not modality or modality == "general_neuromodulation":
                continue
            conditions = split_tags(row.get("indication_tags"))
            for condition in conditions:
                counts[(modality, condition)] += 1
    return [(m, c, n) for (m, c), n in counts.most_common(limit)]


def filter_fda_record_for_modality(rec: dict, modality: str) -> bool:
    haystack = " ".join(
        [
            rec.get("device_name") or "",
            rec.get("generic_name") or "",
            rec.get("applicant") or "",
            rec.get("product_code") or "",
        ]
    ).lower()
    keywords = FDA_DEVICE_KEYWORDS.get(modality) or []
    return any(keyword in haystack for keyword in keywords)


def fetch_trials_and_fda(bundle_root: Path, observed_modalities: set[str]) -> tuple[list[dict], list[dict]]:
    derived = bundle_root / "derived"
    derived.mkdir(exist_ok=True)
    seed_meta = modality_seed_metadata()
    trial_rows: dict[str, dict] = {}
    fda_rows: dict[str, dict] = {}

    combo_queries = top_modality_condition_pairs(bundle_root, limit=36)

    for modality in sorted(observed_modalities):
        seed_modality = AI_TO_SEED_MODALITY.get(modality)
        if not seed_modality:
            continue
        meta = seed_meta.get(seed_modality) or {}
        trial_queries = set(meta.get("trial_queries") or [])
        for combo_modality, condition_slug, _count in combo_queries:
            if combo_modality != modality:
                continue
            for mod_term in MODALITY_QUERY_LABELS.get(modality, [modality.replace("_", " ")]):
                for cond_term in CONDITION_QUERY_LABELS.get(condition_slug, [condition_slug.replace("_", " ")]):
                    trial_queries.add(f'"{mod_term}" "{cond_term}"')
        matched_condition_slugs = sorted(
            {
                condition_slug
                for combo_modality, condition_slug, _count in combo_queries
                if combo_modality == modality
            }
        )
        for trial_query in sorted(trial_queries):
            try:
                studies = ctgov.search(trial_query, max_records=20)
            except Exception:
                studies = []
            for study in studies:
                protocol = study.get("protocolSection", {})
                ident = protocol.get("identificationModule") or {}
                status = protocol.get("statusModule") or {}
                design = protocol.get("designModule") or {}
                nct_id = ident.get("nctId")
                if not nct_id:
                    continue
                trial_rows.setdefault(
                    nct_id,
                    {
                        "nct_id": nct_id,
                        "title": ident.get("briefTitle") or ident.get("officialTitle") or "",
                        "overall_status": status.get("overallStatus") or "",
                        "phase": ";".join(design.get("phases") or []),
                        "study_type": design.get("studyType") or "",
                        "enrollment": str(((design.get("enrollmentInfo") or {}).get("count")) or ""),
                        "conditions_json": json.dumps(((protocol.get("conditionsModule") or {}).get("conditions") or []), ensure_ascii=False),
                        "interventions_json": json.dumps(((protocol.get("armsInterventionsModule") or {}).get("interventions") or []), ensure_ascii=False),
                        "outcomes_json": json.dumps(((protocol.get("outcomesModule") or {}).get("primaryOutcomes") or []), ensure_ascii=False),
                        "brief_summary": ((protocol.get("descriptionModule") or {}).get("briefSummary")) or "",
                        "start_date": ((status.get("startDateStruct") or {}).get("date")) or "",
                        "last_update": ((status.get("lastUpdatePostDateStruct") or {}).get("date")) or "",
                        "sponsor": (((protocol.get("sponsorCollaboratorsModule") or {}).get("leadSponsor") or {}).get("name")) or "",
                        "matched_modality": modality,
                        "matched_conditions": "; ".join(matched_condition_slugs),
                        "query_used": trial_query,
                        "source_url": f"https://clinicaltrials.gov/study/{nct_id}",
                    },
                )
        applicants = sorted(meta.get("applicants") or [])
        product_codes = sorted(meta.get("product_codes") or [])
        for applicant in applicants:
            try:
                records = openfda.search_510k(applicant, max_records=20, product_codes=product_codes if product_codes else None)
            except Exception:
                records = []
            if not product_codes:
                records = [rec for rec in records if filter_fda_record_for_modality(rec, modality)]
            for rec in records:
                number = rec.get("k_number")
                if not number:
                    continue
                fda_rows.setdefault(
                    number,
                    {
                        "clearance_number": number,
                        "decision_date": rec.get("decision_date") or "",
                        "applicant": rec.get("applicant") or "",
                        "device_name": rec.get("device_name") or "",
                        "generic_name": rec.get("generic_name") or "",
                        "product_code": rec.get("product_code") or "",
                        "advisory_committee": rec.get("advisory_committee_description") or "",
                        "matched_modality": modality,
                        "matched_product_codes": "; ".join(product_codes),
                        "query_used": applicant,
                        "source_url": f"https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfpmn/pmn.cfm?ID={number}",
                    },
                )

    trial_list = refine_trial_rows(list(trial_rows.values()))
    fda_list = refine_fda_rows(list(fda_rows.values()))

    write_csv(
        derived / "neuromodulation_clinical_trials.csv",
        [
            "nct_id",
            "title",
            "overall_status",
            "phase",
            "study_type",
            "enrollment",
            "conditions_json",
            "interventions_json",
            "outcomes_json",
            "brief_summary",
            "start_date",
            "last_update",
            "sponsor",
            "matched_modality",
            "matched_conditions",
            "query_used",
            "relevance_score",
            "relevance_bucket",
            "relevance_reasons",
            "source_url",
        ],
        trial_list,
    )
    write_csv(
        derived / "neuromodulation_fda_510k_devices.csv",
        [
            "clearance_number",
            "decision_date",
            "applicant",
            "device_name",
            "generic_name",
            "product_code",
            "advisory_committee",
            "matched_modality",
            "matched_product_codes",
            "query_used",
            "relevance_score",
            "relevance_bucket",
            "relevance_reasons",
            "source_url",
        ],
        fda_list,
    )
    return trial_list, fda_list


def build_research_summary(row: dict, abstract: str, condition_counts: Counter, outcome_rows: list[dict], safety_row: dict | None) -> str:
    parts: list[str] = []
    evidence_tier = row.get("evidence_tier") or "unspecified"
    study_type = row.get("study_type_normalized") or "other"
    modality = row.get("primary_modality") or "unspecified_modality"
    indications = ", ".join(split_tags(row.get("indication_tags"))[:3])
    targets = ", ".join(split_tags(row.get("target_tags"))[:3])
    parameters = ", ".join(split_tags(row.get("parameter_signal_tags"))[:3])
    if evidence_tier or study_type:
        parts.append(f"{study_type} / {evidence_tier} evidence")
    if modality:
        parts.append(f"primary modality: {modality}")
    if indications:
        parts.append(f"indications: {indications}")
    if targets:
        parts.append(f"targets: {targets}")
    if parameters:
        parts.append(f"parameter signals: {parameters}")
    if condition_counts:
        top_conditions = ", ".join(f"{slug} ({count})" for slug, count in condition_counts.most_common(3))
        parts.append(f"condition mentions: {top_conditions}")
    if safety_row:
        safety = normalize_text(safety_row.get("safety_signal_tags"), 120)
        contraindications = normalize_text(safety_row.get("contraindication_signal_tags"), 120)
        if safety:
            parts.append(f"safety: {safety}")
        if contraindications:
            parts.append(f"contraindications: {contraindications}")
    if outcome_rows:
        categories = Counter(r.get("endpoint_category") or "other" for r in outcome_rows)
        parts.append(f"reported outcomes: {', '.join(cat for cat, _ in categories.most_common(3))}")
    abstract_snippet = normalize_text(abstract, 420)
    if abstract_snippet:
        parts.append(f"abstract: {abstract_snippet}")
    return " | ".join(parts)


def build_priority_score(row: dict, abstract_status: str, outcome_count: int, condition_count: int, safety_row: dict | None) -> int:
    score = safe_int(row.get("protocol_relevance_score"))
    score += safe_int(row.get("citation_count")) // 25
    score += 5 if (row.get("open_access_flag") or "").upper() == "Y" else 0
    score += 4 if (row.get("evidence_tier") or "") == "high" else 0
    score += 3 if abstract_status == "ok" else 0
    score += min(outcome_count, 5)
    score += min(condition_count, 5)
    if safety_row and normalize_text(safety_row.get("safety_signal_tags")):
        score += 2
    return score


def build_paper_confidence_score(row: dict) -> int:
    score = 0
    score += {"high": 25, "moderate_high": 20, "moderate": 15, "contextual": 8, "unspecified": 4, "preclinical": 3}.get(
        row.get("evidence_tier") or "", 2
    )
    score += min(safe_int(row.get("cited_by_count")) // 10, 15)
    score += 8 if (row.get("abstract_status") or "") == "ok" else 0
    score += 6 if row.get("journal_normalized") else 0
    score += 6 if row.get("authors_normalized") else 0
    score += min(safe_int(str(row.get("trial_signal_score"))) // 3, 10)
    score += min(safe_int(str(row.get("fda_signal_score"))), 8)
    score += 5 if (row.get("real_world_evidence_flag") or "") == "Y" else 0
    score += 4 if (row.get("is_open_access") or "").upper() == "Y" else 0
    score += min(safe_int(str(row.get("outcome_snippet_count"))), 8)
    return min(score, 100)


def build_enriched_master(bundle_root: Path) -> tuple[list[dict], list[dict], list[dict], set[str]]:
    master_rows = load_csv(bundle_root / "neuromodulation_all_papers_master.csv")
    ai_rows = load_csv(bundle_root / "neuromodulation_ai_ingestion_dataset.csv")
    abstracts = load_csv(bundle_root / "derived" / "neuromodulation_europepmc_abstracts.csv")
    conditions = load_csv(bundle_root / "derived" / "neuromodulation_condition_mentions.csv")
    outcomes = load_csv(bundle_root / "derived" / "neuromodulation_patient_outcomes.csv")
    safety_rows = load_csv(bundle_root / "neuromodulation_safety_contraindication_signals.csv")
    evidence_graph = load_csv(bundle_root / "neuromodulation_evidence_graph.csv")
    protocol_templates = load_csv(bundle_root / "neuromodulation_protocol_template_candidates.csv")

    ai_by_key = {row["paper_key"]: row for row in ai_rows}
    abstract_by_key = {row["paper_key"]: row for row in abstracts}
    safety_by_key = {row["paper_key"]: row for row in safety_rows}
    condition_by_key: dict[str, list[dict]] = defaultdict(list)
    outcomes_by_key: dict[str, list[dict]] = defaultdict(list)
    for row in conditions:
        condition_by_key[row["paper_key"]].append(row)
    for row in outcomes:
        outcomes_by_key[row["paper_key"]].append(row)

    enriched_rows: list[dict] = []
    observed_modalities: set[str] = set()
    for row in master_rows:
        paper_key = ""
        for prefix, field in (("doi", "doi"), ("pmid", "pmid"), ("pmcid", "pmcid")):
            value = (row.get(field) or "").strip()
            if value:
                paper_key = f"{prefix}|{value.lower()}"
                break
        if not paper_key:
            paper_key = f"titleyear|{normalize_text(row.get('title')).lower()}|{row.get('year') or 'unknown'}"

        ai = ai_by_key.get(paper_key, {})
        abstract_row = abstract_by_key.get(paper_key, {})
        safety_row = safety_by_key.get(paper_key)
        abstract = abstract_row.get("abstract") or ""
        condition_counts = Counter()
        for item in condition_by_key.get(paper_key, []):
            condition_counts[item.get("condition_slug") or "unknown"] += safe_int(item.get("match_count"))
        paper_outcomes = outcomes_by_key.get(paper_key, [])

        modality = ai.get("primary_modality") or row.get("matched_modalities") or ""
        if modality:
            observed_modalities.add(modality.split(";")[0].strip())
        journal = normalize_text(ai.get("journal") or row.get("journal"))
        journal_normalized, journal_quality_flag = normalize_journal_name(journal, row.get("source") or ai.get("source"))
        authors = normalize_text(ai.get("authors") or row.get("authors"))
        person_authors, author_groups, author_metadata_notes = separate_group_authors(authors)
        authors_normalized, authors_quality_flag, first_author_normalized, author_count = normalize_authors(person_authors or authors)
        abstract_status = abstract_row.get("retrieval_status") or ("ok" if abstract else "missing")
        outcome_categories = Counter(item.get("endpoint_category") or "other" for item in paper_outcomes)
        top_conditions = "; ".join(f"{slug}:{count}" for slug, count in condition_counts.most_common(5))
        research_summary = build_research_summary(ai or row, abstract, condition_counts, paper_outcomes, safety_row)
        priority_score = build_priority_score(ai or row, abstract_status, len(paper_outcomes), sum(condition_counts.values()), safety_row)

        enriched_rows.append(
            {
                "paper_key": paper_key,
                "priority_score": priority_score,
                "title": row.get("title") or ai.get("title") or "",
                "research_summary": research_summary,
                "abstract": abstract,
                "abstract_status": abstract_status,
                "source": row.get("source") or ai.get("source") or "",
                "id": row.get("id") or "",
                "pmid": row.get("pmid") or ai.get("pmid") or "",
                "pmcid": row.get("pmcid") or ai.get("pmcid") or "",
                "doi": row.get("doi") or ai.get("doi") or "",
                "journal": journal,
                "journal_normalized": journal_normalized,
                "journal_quality_flag": journal_quality_flag,
                "journal_backfill_source": "",
                "authors": authors,
                "authors_normalized": authors_normalized,
                "authors_quality_flag": authors_quality_flag,
                "first_author_normalized": first_author_normalized,
                "author_groups": author_groups,
                "author_metadata_notes": author_metadata_notes,
                "author_group_flag": "Y" if author_groups else "N",
                "author_count": author_count,
                "year": row.get("year") or ai.get("year") or "",
                "first_publication_date": row.get("first_publication_date") or "",
                "first_index_date": row.get("first_index_date") or "",
                "pub_type": row.get("pub_type") or ai.get("pub_type_raw") or "",
                "study_type_normalized": ai.get("study_type_normalized") or "",
                "evidence_tier": ai.get("evidence_tier") or "",
                "primary_modality": ai.get("primary_modality") or "",
                "canonical_modalities": ai.get("canonical_modalities") or row.get("matched_modalities") or "",
                "invasiveness": ai.get("invasiveness") or "",
                "indication_tags": ai.get("indication_tags") or "",
                "condition_mentions_top": top_conditions,
                "population_tags": ai.get("population_tags") or "",
                "target_tags": ai.get("target_tags") or "",
                "parameter_signal_tags": ai.get("parameter_signal_tags") or "",
                "safety_signal_tags": (safety_row or {}).get("safety_signal_tags") or "",
                "contraindication_signal_tags": (safety_row or {}).get("contraindication_signal_tags") or "",
                "protocol_relevance_score": ai.get("protocol_relevance_score") or "",
                "matched_query_terms": row.get("matched_query_terms") or ai.get("matched_query_terms") or "",
                "source_exports": row.get("source_exports") or ai.get("source_exports") or "",
                "is_open_access": row.get("is_open_access") or ai.get("open_access_flag") or "",
                "cited_by_count": row.get("cited_by_count") or ai.get("citation_count") or "",
                "outcome_snippet_count": len(paper_outcomes),
                "outcome_categories": "; ".join(cat for cat, _ in outcome_categories.most_common(5)),
                "real_world_evidence_flag": "Y" if any((item.get("real_world_evidence_flag") or "").upper() == "Y" for item in paper_outcomes) else "N",
                "record_url": ai.get("record_url") or row.get("europe_pmc_url") or "",
                "ai_ingestion_text": ai.get("ai_ingestion_text") or "",
            }
        )
        enriched_rows[-1]["paper_confidence_score"] = build_paper_confidence_score(enriched_rows[-1])
    return enriched_rows, evidence_graph, protocol_templates, observed_modalities


def write_overview_sheet(ws, overview_rows: list[tuple[str, str]]) -> None:
    ws.title = "Overview"
    ws.append(["Metric", "Value"])
    for key, value in overview_rows:
        ws.append([key, value])
    style_sheet(ws)


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = CELL_BORDER
        cell.alignment = WRAP
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = CELL_BORDER
            cell.alignment = WRAP
    for col_idx, column in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column[:600]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)


def append_sheet(wb: Workbook, title: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["No data"])
        return
    fieldnames = list(rows[0].keys())
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(field, "") for field in fieldnames])
    style_sheet(ws)


def build_workbook(
    out_path: Path,
    overview_rows: list[tuple[str, str]],
    enriched_rows: list[dict],
    evidence_graph: list[dict],
    protocol_templates: list[dict],
    trial_rows: list[dict],
    fda_rows: list[dict],
    top_conditions_rows: list[dict],
    outcome_rows: list[dict],
) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    overview = wb.create_sheet("Overview")
    write_overview_sheet(overview, overview_rows)
    append_sheet(wb, "Enriched_Master", enriched_rows)
    append_sheet(wb, "Top_Conditions", top_conditions_rows)
    append_sheet(wb, "Patient_Outcomes", outcome_rows)
    append_sheet(wb, "Clinical_Trials", trial_rows)
    append_sheet(wb, "FDA_510k", fda_rows)
    append_sheet(wb, "Protocol_Templates", protocol_templates[:20000])
    append_sheet(wb, "Evidence_Graph", evidence_graph[:50000])
    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--bundle-root", type=Path, required=True)
    parser.add_argument("--journal-backfill-limit", type=int, default=0)
    parser.add_argument("--journal-backfill-min-priority", type=int, default=30)
    parser.add_argument("--journal-backfill-min-confidence", type=int, default=25)
    parser.add_argument("--skip-xlsx", action="store_true")
    args = parser.parse_args()

    bundle_root = args.bundle_root
    enriched_rows, evidence_graph, protocol_templates, observed_modalities = build_enriched_master(bundle_root)
    derived = bundle_root / "derived"
    trial_rows = []
    fda_rows = []
    trials_path = derived / "neuromodulation_clinical_trials.csv"
    fda_path = derived / "neuromodulation_fda_510k_devices.csv"
    if trials_path.exists():
        trial_rows = load_csv(trials_path)
    if fda_path.exists():
        fda_rows = load_csv(fda_path)
    if not trial_rows and not fda_rows:
        trial_rows, fda_rows = fetch_trials_and_fda(bundle_root, observed_modalities)
    else:
        trial_rows = refine_trial_rows(trial_rows)
        fda_rows = refine_fda_rows(fda_rows)
        if trial_rows:
            write_csv(
                trials_path,
                [
                    "nct_id",
                    "title",
                    "overall_status",
                    "phase",
                    "study_type",
                    "enrollment",
                    "conditions_json",
                    "interventions_json",
                    "outcomes_json",
                    "brief_summary",
                    "start_date",
                    "last_update",
                    "sponsor",
                    "matched_modality",
                    "matched_conditions",
                    "query_used",
                    "relevance_score",
                    "relevance_bucket",
                    "relevance_reasons",
                    "protocol_frequency_hz",
                    "protocol_pulse_width_us",
                    "protocol_current_ma",
                    "protocol_voltage_v",
                    "protocol_duration",
                    "protocol_session_count",
                    "protocol_pulse_count",
                    "protocol_duty_cycle",
                    "protocol_target_site",
                    "protocol_has_sham_control",
                    "protocol_parameter_summary",
                    "source_url",
                ],
                trial_rows,
            )
        if fda_rows:
            write_csv(
                fda_path,
                [
                    "clearance_number",
                    "decision_date",
                    "applicant",
                    "device_name",
                    "generic_name",
                    "product_code",
                    "advisory_committee",
                    "matched_modality",
                    "matched_product_codes",
                    "query_used",
                    "relevance_score",
                    "relevance_bucket",
                    "relevance_reasons",
                    "source_url",
                ],
                fda_rows,
            )

    attach_paper_level_external_signals(enriched_rows, trial_rows, fda_rows)
    journal_backfill_stats = {"candidate_count": 0, "attempted": 0, "updated": 0}
    if args.journal_backfill_limit > 0:
        journal_backfill_stats = apply_bounded_journal_backfill(
            enriched_rows,
            derived,
            limit=args.journal_backfill_limit,
            min_priority_score=args.journal_backfill_min_priority,
            min_confidence_score=args.journal_backfill_min_confidence,
        )
    for row in enriched_rows:
        row["paper_confidence_score"] = build_paper_confidence_score(row)

    condition_rows = load_csv(bundle_root / "derived" / "neuromodulation_condition_mentions.csv")
    outcome_rows = load_csv(bundle_root / "derived" / "neuromodulation_patient_outcomes.csv")
    condition_totals: Counter[str] = Counter()
    for row in condition_rows:
        condition_totals[row.get("condition_slug") or "unknown"] += safe_int(row.get("match_count"))
    top_conditions_rows = [
        {"condition_slug": slug, "mention_count": count}
        for slug, count in condition_totals.most_common(250)
    ]

    enriched_rows.sort(
        key=lambda row: (
            safe_int(str(row.get("priority_score"))),
            safe_int(str(row.get("cited_by_count"))),
            safe_int(str(row.get("year"))),
        ),
        reverse=True,
    )
    overview_rows = [
        ("Bundle root", str(bundle_root)),
        ("Enriched master rows", str(len(enriched_rows))),
        ("Rows with abstract text", str(sum(1 for row in enriched_rows if row.get("abstract")))),
        ("Rows with journal name", str(sum(1 for row in enriched_rows if row.get("journal")))),
        ("Rows with authors", str(sum(1 for row in enriched_rows if row.get("authors")))),
        ("Rows with condition mentions", str(sum(1 for row in enriched_rows if row.get("condition_mentions_top")))),
        ("Rows with outcome snippets", str(sum(1 for row in enriched_rows if safe_int(str(row.get("outcome_snippet_count")))))),
        ("Real-world evidence flagged rows", str(sum(1 for row in enriched_rows if row.get("real_world_evidence_flag") == "Y"))),
        ("Distinct primary modalities", str(len({row.get('primary_modality') for row in enriched_rows if row.get('primary_modality')}))),
        ("ClinicalTrials.gov rows", str(len(trial_rows))),
        ("FDA 510(k) rows", str(len(fda_rows))),
        ("Journal backfill attempted", str(journal_backfill_stats["attempted"])),
        ("Journal backfill updated", str(journal_backfill_stats["updated"])),
        ("Mean priority score", f"{mean([safe_int(str(row.get('priority_score'))) for row in enriched_rows[:5000]]) if enriched_rows else 0:.2f}"),
    ]

    enriched_csv = bundle_root / "neuromodulation_master_database_enriched.csv"
    enriched_xlsx = bundle_root / "deepsynaps-evidence-enriched.xlsx"
    write_csv(enriched_csv, list(enriched_rows[0].keys()), enriched_rows)
    if not args.skip_xlsx:
        build_workbook(
            enriched_xlsx,
            overview_rows,
            enriched_rows,
            evidence_graph,
            protocol_templates,
            trial_rows,
            fda_rows,
            top_conditions_rows,
            outcome_rows,
        )
    print(
        json.dumps(
            {
                "enriched_csv": str(enriched_csv),
                "enriched_xlsx": "" if args.skip_xlsx else str(enriched_xlsx),
                "master_rows": len(enriched_rows),
                "trial_rows": len(trial_rows),
                "fda_rows": len(fda_rows),
                "journal_backfill": journal_backfill_stats,
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
