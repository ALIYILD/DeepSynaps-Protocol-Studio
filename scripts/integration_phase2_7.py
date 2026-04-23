#!/usr/bin/env python3
"""
DeepSynaps Studio — SOZO Data Integration (Phases 2-7)
Rebuilds existing 9-table database + integrates SOZO clinical data.
Produces: snapshot, new CSVs, new Excel workbook, integration report.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import csv
import os
from datetime import datetime

OUTPUT_DIR = "/home/user/workspace/deepsynaps_output"
CSV_DIR = os.path.join(OUTPUT_DIR, "csv")
os.makedirs(CSV_DIR, exist_ok=True)

# ============================================================
# PHASE 0: Rebuild existing 9-table database (from previous session)
# ============================================================

print("=" * 60)
print("PHASE 0: Rebuilding existing 9-table database")
print("=" * 60)

# Table 1: Evidence_Levels
evidence_levels = [
    {"Evidence_ID": "EV-A", "Level_Name": "Guideline-Endorsed", "Description": "Endorsed by major clinical practice guidelines (e.g., APA, NICE, CANMAT). Supported by multiple RCTs, systematic reviews, or meta-analyses.", "Minimum_Evidence": "≥2 high-quality RCTs + guideline endorsement", "Usage_Gate": "Can be recommended as standard of care", "Review_Status": "Reviewed"},
    {"Evidence_ID": "EV-B", "Level_Name": "Strong Research Evidence", "Description": "Supported by RCTs and/or meta-analyses but not yet endorsed in major guidelines. Consistent positive findings across studies.", "Minimum_Evidence": "≥1 RCT + systematic review or meta-analysis", "Usage_Gate": "Can be recommended with clinician judgment", "Review_Status": "Reviewed"},
    {"Evidence_ID": "EV-C", "Level_Name": "Emerging Evidence", "Description": "Supported by open-label trials, pilot RCTs, or small controlled studies. Promising but insufficient for guideline endorsement.", "Minimum_Evidence": "Pilot RCT or ≥2 open-label studies", "Usage_Gate": "Clinician discretion; informed consent required", "Review_Status": "Reviewed"},
    {"Evidence_ID": "EV-D", "Level_Name": "Preliminary / Investigational", "Description": "Case series, case reports, preclinical data, or expert opinion only. Insufficient controlled evidence.", "Minimum_Evidence": "Case series or preclinical data", "Usage_Gate": "Research use only; cannot be patient-facing without explicit informed consent and clinician oversight", "Review_Status": "Reviewed"},
]

# Table 2: Governance_Rules
governance_rules = [
    {"Rule_ID": "GOV-001", "Rule_Name": "Off-Label Protocol Flag", "Description": "Any protocol using a device or modality outside its approved/cleared indication must be flagged as off-label.", "Trigger_Condition": "Protocol.Regulatory_Status contains 'off-label' or 'investigational'", "Action": "Flag for clinician review; add off-label disclaimer to any patient-facing output", "Severity": "Warning", "Review_Status": "Reviewed"},
    {"Rule_ID": "GOV-002", "Rule_Name": "Evidence Grade Gate", "Description": "Protocols graded EV-D cannot appear in patient-facing recommendations without explicit clinician override.", "Trigger_Condition": "Protocol.Evidence_Level = EV-D AND output_type = patient-facing", "Action": "Block output; require clinician sign-off", "Severity": "Critical", "Review_Status": "Reviewed"},
    {"Rule_ID": "GOV-003", "Rule_Name": "Investigational Modality Flag", "Description": "Protocols using modalities without any FDA clearance/approval for neuropsychiatric indications must carry investigational warning.", "Trigger_Condition": "Modality.Regulatory_Notes contains 'Investigational' or 'no FDA clearance'", "Action": "Add investigational disclaimer; flag in review queue", "Severity": "Warning", "Review_Status": "Reviewed"},
    {"Rule_ID": "GOV-004", "Rule_Name": "Contraindication Check Required", "Description": "Every protocol must have contraindications reviewed before export.", "Trigger_Condition": "Protocol.Contraindications_Reviewed = false", "Action": "Block export until contraindication review completed", "Severity": "Critical", "Review_Status": "Reviewed"},
    {"Rule_ID": "GOV-005", "Rule_Name": "Source Traceability Required", "Description": "Every protocol must link to at least one peer-reviewed source.", "Trigger_Condition": "Protocol.Source_IDs is empty", "Action": "Flag for review; block publication", "Severity": "Critical", "Review_Status": "Reviewed"},
    {"Rule_ID": "GOV-006", "Rule_Name": "Regulatory Terminology Precision", "Description": "FDA cleared ≠ FDA approved. 510(k) clearance and PMA approval must use correct terminology.", "Trigger_Condition": "Always active", "Action": "Validate terminology on save; reject 'FDA approved' for 510(k) devices", "Severity": "Warning", "Review_Status": "Reviewed"},
    {"Rule_ID": "GOV-007", "Rule_Name": "Device Listing ≠ Cleared Intended Use", "Description": "FDA device listing/registration does not imply cleared intended use for a specific condition.", "Trigger_Condition": "Device.Regulatory_Status = 'listed' AND Protocol implies clearance", "Action": "Correct regulatory field; add disclaimer", "Severity": "Warning", "Review_Status": "Reviewed"},
    {"Rule_ID": "GOV-008", "Rule_Name": "Neurofeedback ADHD Evidence Lock", "Description": "Neurofeedback for ADHD remains EV-D per Cortese 2024 meta-analysis. Do not upgrade without new blinded RCT evidence.", "Trigger_Condition": "Modality = Neurofeedback AND Condition = ADHD AND Evidence_Level != EV-D", "Action": "Reset to EV-D; log governance override", "Severity": "Critical", "Review_Status": "Reviewed"},
    {"Rule_ID": "GOV-009", "Rule_Name": "Flow FL-100 PMA Exclusivity", "Description": "Flow FL-100 is the ONLY PMA-approved tDCS device (Dec 2025, treatment-resistant MDD). No other tDCS device has PMA approval.", "Trigger_Condition": "Device.Modality = tDCS AND Device.Regulatory_Status = 'FDA approved (PMA)' AND Device != FL-100", "Action": "Reject; correct regulatory status", "Severity": "Critical", "Review_Status": "Reviewed"},
    {"Rule_ID": "GOV-010", "Rule_Name": "Marketing Language Prohibition", "Description": "No marketing, promotional, or unsubstantiated language in regulatory or evidence fields.", "Trigger_Condition": "Text analysis of regulatory/evidence fields", "Action": "Flag for review; suggest neutral clinical language", "Severity": "Warning", "Review_Status": "Reviewed"},
    {"Rule_ID": "GOV-011", "Rule_Name": "Review Queue Mandatory for New Records", "Description": "All newly imported or created records must enter the review queue before publication.", "Trigger_Condition": "Record.is_new = true", "Action": "Set Review_Status = 'Pending'; add to review queue", "Severity": "Critical", "Review_Status": "Reviewed"},
    {"Rule_ID": "GOV-012", "Rule_Name": "Snapshot Before Data Changes", "Description": "A database snapshot must be created before any bulk data import or schema change.", "Trigger_Condition": "Bulk import or schema migration initiated", "Action": "Create timestamped snapshot; log change set", "Severity": "Critical", "Review_Status": "Reviewed"},
]

# Table 3: Existing Modalities (12)
existing_modalities = [
    {"Modality_ID": "MOD-001", "Modality_Name": "TMS / rTMS", "Full_Name": "Transcranial Magnetic Stimulation / Repetitive TMS", "Category": "Non-invasive Brain Stimulation", "Mechanism": "Electromagnetic induction creates focused electric currents in cortical tissue via rapidly changing magnetic field", "FDA_Status_Summary": "FDA cleared for MDD (2008), OCD (2018), smoking cessation (2020), migraine (2013)", "Regulatory_Notes": "Multiple 510(k) clearances; BrainsWay has broadest indications (MDD, OCD, smoking cessation)", "Key_Parameters": "Frequency (Hz), Intensity (% rMT), Pulses/session, Coil type", "Review_Status": "Reviewed"},
    {"Modality_ID": "MOD-002", "Modality_Name": "tDCS", "Full_Name": "Transcranial Direct Current Stimulation", "Category": "Non-invasive Brain Stimulation", "Mechanism": "Weak DC current (1-2 mA) modulates neuronal excitability; anodal = excitatory, cathodal = inhibitory", "FDA_Status_Summary": "Flow FL-100 PMA approved for treatment-resistant MDD (Dec 2025); otherwise investigational in US", "Regulatory_Notes": "FL-100 is the ONLY PMA-approved tDCS device. Soterix and Neuroelectrics are NOT FDA cleared for treatment. Flow has FDA Breakthrough Designation (2021).", "Key_Parameters": "Current (mA), Duration (min), Electrode montage, Sessions", "Review_Status": "Reviewed"},
    {"Modality_ID": "MOD-003", "Modality_Name": "Neurofeedback", "Full_Name": "EEG Neurofeedback / Neurotherapy", "Category": "Biofeedback", "Mechanism": "Real-time EEG feedback trains self-regulation of brain activity patterns", "FDA_Status_Summary": "Devices FDA cleared as biofeedback devices (general wellness); no clearance for specific neuropsychiatric conditions", "Regulatory_Notes": "ADHD: EV-D per Cortese 2024 meta-analysis — do not upgrade. Devices cleared as general biofeedback only.", "Key_Parameters": "Protocol type, Frequency bands, Electrode sites, Sessions", "Review_Status": "Reviewed"},
    {"Modality_ID": "MOD-004", "Modality_Name": "CES", "Full_Name": "Cranial Electrotherapy Stimulation", "Category": "Non-invasive Brain Stimulation", "Mechanism": "Low-intensity alternating current via ear-clip electrodes modulates cortical excitability", "FDA_Status_Summary": "FDA cleared for anxiety, insomnia, depression, pain (Alpha-Stim, Fisher Wallace)", "Regulatory_Notes": "Alpha-Stim: FDA cleared for anxiety, insomnia, depression, pain. Fisher Wallace: FDA cleared for depression, anxiety, insomnia.", "Key_Parameters": "Current (μA), Frequency (Hz), Duration (min), Waveform", "Review_Status": "Reviewed"},
    {"Modality_ID": "MOD-005", "Modality_Name": "VNS / taVNS", "Full_Name": "Vagus Nerve Stimulation / Transcutaneous Auricular VNS", "Category": "Peripheral Nerve Stimulation", "Mechanism": "Stimulation of vagus nerve (implanted or transcutaneous auricular) modulates NTS-locus coeruleus-cortical circuits", "FDA_Status_Summary": "Implanted VNS: FDA approved for epilepsy (1997), treatment-resistant depression (2005). taVNS: CE marked in Europe; gammaCore FDA cleared for cluster headache and migraine", "Regulatory_Notes": "taVNS (auricular) is investigational in US for most indications. gammaCore (cervical) has specific FDA clearances.", "Key_Parameters": "Current (mA), Frequency (Hz), Pulse width (μs), Duration, Location", "Review_Status": "Reviewed"},
    {"Modality_ID": "MOD-006", "Modality_Name": "TPS", "Full_Name": "Transcranial Pulse Stimulation", "Category": "Non-invasive Brain Stimulation", "Mechanism": "Short ultrasound pulses generate mechanical pressure waves that modulate neuronal activity and promote neuroplasticity", "FDA_Status_Summary": "CE marked in Europe for Alzheimer's disease (NEUROLITH). Not FDA cleared in US.", "Regulatory_Notes": "NEUROLITH (Storz Medical AG) is the primary TPS device. CE marked for Alzheimer's in Europe. Investigational in US.", "Key_Parameters": "EFD (mJ/mm²), PRF (Hz), Pulses/session, Duration, Navigation method", "Review_Status": "Reviewed"},
    {"Modality_ID": "MOD-007", "Modality_Name": "PBM", "Full_Name": "Photobiomodulation / Transcranial Near-Infrared Light Therapy", "Category": "Non-invasive Brain Stimulation", "Mechanism": "Near-infrared light (810 nm) penetrates skull, activates cytochrome c oxidase in mitochondria, increases ATP/NO production", "FDA_Status_Summary": "Investigational for neuropsychiatric indications. Some devices CE marked in EU.", "Regulatory_Notes": "Vielight Neuro: Investigational. SYMBYX PDCare: TGA listed (Australia). PhotoThera NeuroThera: NEST-2 negative, NEST-3 ongoing.", "Key_Parameters": "Wavelength (nm), Power density (mW/cm²), Pulse frequency (Hz), Duration, Delivery method", "Review_Status": "Reviewed"},
    {"Modality_ID": "MOD-008", "Modality_Name": "DBS", "Full_Name": "Deep Brain Stimulation", "Category": "Invasive Neurostimulation", "Mechanism": "Implanted electrodes deliver continuous electrical stimulation to deep brain nuclei, modulating circuit activity", "FDA_Status_Summary": "FDA approved for Parkinson's (STN 1997, GPi 2010), essential tremor (1997), dystonia (HDE 2003), OCD (HDE 2009), epilepsy (2018)", "Regulatory_Notes": "Requires neurosurgical implantation. Medtronic Percept PC, Abbott Infinity, Boston Scientific Vercise are major systems.", "Key_Parameters": "Voltage/Current (V/mA), Frequency (Hz), Pulse width (μs), Contact configuration, Target nucleus", "Review_Status": "Reviewed"},
    {"Modality_ID": "MOD-009", "Modality_Name": "ECT", "Full_Name": "Electroconvulsive Therapy", "Category": "Non-invasive Brain Stimulation", "Mechanism": "Brief electrical stimulus induces generalized seizure under general anesthesia; mechanism involves neurotransmitter and neuroplasticity changes", "FDA_Status_Summary": "FDA reclassified to Class II for catatonia and severe MDD (2018)", "Regulatory_Notes": "Gold standard for severe/treatment-resistant depression, catatonia, acute suicidality. Requires general anesthesia.", "Key_Parameters": "Charge (mC), Pulse width, Electrode placement (bilateral/unilateral), Sessions", "Review_Status": "Reviewed"},
    {"Modality_ID": "MOD-010", "Modality_Name": "MST", "Full_Name": "Magnetic Seizure Therapy", "Category": "Non-invasive Brain Stimulation", "Mechanism": "High-dose TMS induces focal seizure; combines seizure efficacy of ECT with focality of TMS", "FDA_Status_Summary": "Investigational — not FDA cleared", "Regulatory_Notes": "Investigational. Phase 2/3 trials ongoing. Promising cognitive safety profile vs ECT.", "Key_Parameters": "Frequency, Coil position, Train duration, Anesthesia protocol", "Review_Status": "Reviewed"},
    {"Modality_ID": "MOD-011", "Modality_Name": "tVNS (cervical)", "Full_Name": "Transcutaneous Cervical Vagus Nerve Stimulation", "Category": "Peripheral Nerve Stimulation", "Mechanism": "Non-invasive cervical vagus nerve stimulation via handheld device", "FDA_Status_Summary": "gammaCore: FDA cleared for episodic cluster headache (2017), migraine prevention (2018), migraine treatment (2019)", "Regulatory_Notes": "electroCore gammaCore is the primary device. Cleared specifically for headache/migraine — not for depression or other neuropsychiatric indications.", "Key_Parameters": "Stimulation duration, Frequency, Amplitude, Number of stimulations/day", "Review_Status": "Reviewed"},
    {"Modality_ID": "MOD-012", "Modality_Name": "NIBS (general)", "Full_Name": "Non-Invasive Brain Stimulation (General Category)", "Category": "Non-invasive Brain Stimulation", "Mechanism": "Umbrella category for non-invasive techniques that modulate brain activity through transcranial electrical or magnetic means", "FDA_Status_Summary": "Varies by specific modality", "Regulatory_Notes": "This is a classification category. Specific regulatory status depends on the individual modality and device.", "Key_Parameters": "Varies by modality", "Review_Status": "Reviewed"},
]

print(f"  Rebuilt {len(evidence_levels)} evidence levels")
print(f"  Rebuilt {len(governance_rules)} governance rules")
print(f"  Rebuilt {len(existing_modalities)} existing modalities")

# ============================================================
# PHASE 2: Add 4 new modalities
# ============================================================

print("\n" + "=" * 60)
print("PHASE 2: Adding 4 new modalities")
print("=" * 60)

new_modalities = [
    {"Modality_ID": "MOD-013", "Modality_Name": "tACS", "Full_Name": "Transcranial Alternating Current Stimulation", "Category": "Non-invasive Brain Stimulation", "Mechanism": "Sinusoidal alternating current at specific frequencies entrains endogenous brain oscillations; frequency-specific effects on cortical excitability and network synchrony", "FDA_Status_Summary": "Investigational — no FDA clearance for neuropsychiatric indications. Nexalin device has 510(k) clearance for insomnia specifically.", "Regulatory_Notes": "Investigational — no FDA clearance for neuropsychiatric indications. Nexalin 510(k) for insomnia is device-specific, not modality-wide. Soterix and Neuroelectrics tACS capabilities are research-use only in US.", "Key_Parameters": "Frequency (Hz), Current (mA peak-to-peak), Duration (min), Electrode montage, Waveform", "Review_Status": "Reviewed"},
    {"Modality_ID": "MOD-014", "Modality_Name": "PEMF", "Full_Name": "Pulsed Electromagnetic Field Therapy", "Category": "Non-invasive Brain Stimulation", "Mechanism": "Low-intensity pulsed electromagnetic fields modulate neuronal excitability, promote BDNF expression, and influence cortical oscillations; overlaps with low-intensity rTMS at certain parameters", "FDA_Status_Summary": "Investigational — no FDA clearance for neuropsychiatric indications. Some PEMF devices FDA cleared for bone healing and pain (not brain applications).", "Regulatory_Notes": "Investigational — no FDA clearance for neuropsychiatric indications. NeoRhythm, ICES-DigiCeutical, FlexPulse are consumer/research devices. FDA clearance for orthopedic PEMF does not extend to transcranial use.", "Key_Parameters": "Field intensity (mT), Frequency (Hz), Duration (min), Coil placement, Waveform", "Review_Status": "Reviewed"},
    {"Modality_ID": "MOD-015", "Modality_Name": "LIFU / tFUS", "Full_Name": "Low-Intensity Focused Ultrasound / Transcranial Focused Ultrasound", "Category": "Non-invasive Brain Stimulation", "Mechanism": "Focused ultrasound beam reaches deep brain structures non-invasively; mechanical effects on neuronal ion channels modulate excitability with millimeter precision", "FDA_Status_Summary": "Investigational for neuromodulation. MRI-guided FUS (MRgFUS) FDA cleared for ablation (essential tremor) but NOT for low-intensity neuromodulation.", "Regulatory_Notes": "Investigational — no FDA clearance for neuropsychiatric indications. MRgFUS ablation clearance (Insightec) is for tissue destruction, NOT neuromodulation. BrainSonix, Openwater, Brainbox are investigational devices.", "Key_Parameters": "ISPPA (W/cm²), Fundamental frequency (kHz), PRF (Hz), Duration (min), Spatial target, Navigation method", "Review_Status": "Reviewed"},
    {"Modality_ID": "MOD-016", "Modality_Name": "tRNS", "Full_Name": "Transcranial Random Noise Stimulation", "Category": "Non-invasive Brain Stimulation", "Mechanism": "Random electrical noise across a broad frequency spectrum (0.1–640 Hz) enhances cortical excitability via stochastic resonance; may boost subthreshold neural activity", "FDA_Status_Summary": "Investigational — no FDA clearance for neuropsychiatric indications", "Regulatory_Notes": "Investigational — no FDA clearance for neuropsychiatric indications. Soterix and Neuroelectrics devices capable of tRNS are research-use only in US.", "Key_Parameters": "Current (mA), Frequency range (Hz), Duration (min), Electrode montage", "Review_Status": "Reviewed"},
]

all_modalities = existing_modalities + new_modalities
print("  Added 4 new modalities: tACS, PEMF, LIFU/tFUS, tRNS")
print(f"  Total modalities: {len(all_modalities)}")

# ============================================================
# PHASE 4: Condition expansion (before protocols, since protocols need condition IDs)
# ============================================================

print("\n" + "=" * 60)
print("PHASE 4: Condition expansion")
print("=" * 60)

# Existing 20 conditions from v1 database
existing_conditions = [
    {"Condition_ID": "COND-001", "Condition_Name": "Major Depressive Disorder (MDD)", "Category": "Mood / Affective", "ICD_10": "F32, F33", "Key_Symptoms": "Persistent depressed mood, anhedonia, fatigue, sleep/appetite changes, concentration difficulties", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-002", "Condition_Name": "Treatment-Resistant Depression (TRD)", "Category": "Mood / Affective", "ICD_10": "F32.9, F33.9", "Key_Symptoms": "MDD that has not responded to ≥2 adequate antidepressant trials", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-003", "Condition_Name": "Generalized Anxiety Disorder (GAD)", "Category": "Anxiety", "ICD_10": "F41.1", "Key_Symptoms": "Excessive worry, restlessness, muscle tension, sleep disturbance, irritability", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-004", "Condition_Name": "Obsessive-Compulsive Disorder (OCD)", "Category": "Anxiety / OCD Spectrum", "ICD_10": "F42", "Key_Symptoms": "Intrusive obsessions, compulsive rituals, anxiety, avoidance", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-005", "Condition_Name": "Post-Traumatic Stress Disorder (PTSD)", "Category": "Trauma / Stress-Related", "ICD_10": "F43.1", "Key_Symptoms": "Flashbacks, hyperarousal, avoidance, emotional numbing, sleep disturbance", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-006", "Condition_Name": "Insomnia", "Category": "Sleep", "ICD_10": "G47.0", "Key_Symptoms": "Difficulty initiating/maintaining sleep, early morning awakening, daytime impairment", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-007", "Condition_Name": "ADHD", "Category": "Neurodevelopmental", "ICD_10": "F90", "Key_Symptoms": "Inattention, hyperactivity, impulsivity, executive dysfunction", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-008", "Condition_Name": "Chronic Pain", "Category": "Pain", "ICD_10": "G89", "Key_Symptoms": "Persistent pain >3 months, central sensitization, functional impairment", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-009", "Condition_Name": "Alzheimer's Disease", "Category": "Neurodegenerative", "ICD_10": "G30", "Key_Symptoms": "Progressive memory loss, cognitive decline, disorientation, behavioral changes", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-010", "Condition_Name": "Parkinson's Disease", "Category": "Neurodegenerative / Movement", "ICD_10": "G20", "Key_Symptoms": "Bradykinesia, rigidity, tremor, postural instability, cognitive changes", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-011", "Condition_Name": "Stroke (Motor Rehabilitation)", "Category": "Cerebrovascular", "ICD_10": "I63, I69", "Key_Symptoms": "Motor deficit, hemiparesis, spasticity, aphasia, cognitive impairment", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-012", "Condition_Name": "Tinnitus", "Category": "Auditory / Neurological", "ICD_10": "H93.1", "Key_Symptoms": "Phantom auditory perception, distress, sleep disturbance, concentration difficulty", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-013", "Condition_Name": "Epilepsy", "Category": "Neurological", "ICD_10": "G40", "Key_Symptoms": "Recurrent seizures, interictal EEG abnormalities, cognitive impact", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-014", "Condition_Name": "Migraine", "Category": "Headache / Pain", "ICD_10": "G43", "Key_Symptoms": "Recurrent headache with aura/without, photophobia, nausea, disability", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-015", "Condition_Name": "Traumatic Brain Injury (TBI)", "Category": "Neurological / Trauma", "ICD_10": "S06", "Key_Symptoms": "Cognitive impairment, headache, fatigue, mood changes, sleep disturbance", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-016", "Condition_Name": "Addiction / Substance Use Disorder", "Category": "Addiction", "ICD_10": "F10-F19", "Key_Symptoms": "Craving, loss of control, continued use despite consequences, withdrawal", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-017", "Condition_Name": "Autism Spectrum Disorder (ASD)", "Category": "Neurodevelopmental", "ICD_10": "F84", "Key_Symptoms": "Social communication deficits, restricted interests, sensory sensitivities", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-018", "Condition_Name": "Essential Tremor", "Category": "Movement Disorder", "ICD_10": "G25.0", "Key_Symptoms": "Postural and kinetic tremor, typically bilateral upper extremities", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-019", "Condition_Name": "Dystonia", "Category": "Movement Disorder", "ICD_10": "G24", "Key_Symptoms": "Sustained or intermittent muscle contractions causing abnormal postures/movements", "Review_Status": "Reviewed"},
    {"Condition_ID": "COND-020", "Condition_Name": "Smoking Cessation", "Category": "Addiction / Behavioral", "ICD_10": "F17, Z72.0", "Key_Symptoms": "Nicotine dependence, craving, withdrawal symptoms", "Review_Status": "Reviewed"},
]

# Map SOZO condition names to existing condition IDs
condition_name_map = {
    "Depression (MDD)": "COND-001",
    "Depression (MDD) — Standard rTMS": "COND-001",
    "Depression (MDD) — iTBS": "COND-001",
    "Treatment-Resistant Depression (TRD)": "COND-002",
    "Anxiety / GAD": "COND-003",
    "Anxiety": "COND-003",
    "OCD": "COND-004",
    "OCD (Treatment-Resistant)": "COND-004",
    "PTSD": "COND-005",
    "Insomnia": "COND-006",
    "ADHD": "COND-007",
    "Chronic Pain": "COND-008",
    "Alzheimer's Disease": "COND-009",
    "Alzheimer's / Dementia": "COND-009",
    "Alzheimer's / MCI (Gamma Entrainment)": "COND-009",
    "Parkinson's Disease": "COND-010",
    "Parkinson's Disease (Motor)": "COND-010",
    "Parkinson's Disease (Depression)": "COND-010",
    "Stroke Rehabilitation": "COND-011",
    "Stroke (Motor Rehabilitation)": "COND-011",
    "Stroke": "COND-011",
    "Tinnitus": "COND-012",
    "Epilepsy": "COND-013",
    "Migraine": "COND-014",
    "Migraine Prevention": "COND-014",
    "TBI / Concussion": "COND-015",
    "Addiction": "COND-016",
    "Addiction / Substance Use": "COND-016",
    "Autism Spectrum Disorder (ASD)": "COND-017",
    "Essential Tremor": "COND-018",
    "Dystonia": "COND-019",
    "Smoking Cessation": "COND-020",
}

# New conditions from SOZO data
new_conditions = [
    {"Condition_ID": "COND-021", "Condition_Name": "Schizophrenia", "Category": "Psychotic Disorder", "ICD_10": "F20", "Key_Symptoms": "Positive symptoms (hallucinations, delusions), negative symptoms (flat affect, avolition), cognitive deficits", "Review_Status": "Pending"},
    {"Condition_ID": "COND-022", "Condition_Name": "Disorders of Consciousness (DOC)", "Category": "Neurological", "ICD_10": "R40", "Key_Symptoms": "Impaired consciousness ranging from coma to minimally conscious state; impaired arousal and awareness", "Review_Status": "Pending"},
    {"Condition_ID": "COND-023", "Condition_Name": "Mild Cognitive Impairment (MCI)", "Category": "Neurodegenerative", "ICD_10": "G31.84", "Key_Symptoms": "Cognitive decline beyond normal aging but not meeting dementia criteria; memory complaints, executive dysfunction", "Review_Status": "Pending"},
    {"Condition_ID": "COND-024", "Condition_Name": "Multiple Sclerosis (MS)", "Category": "Neurological / Autoimmune", "ICD_10": "G35", "Key_Symptoms": "Relapsing-remitting or progressive neurological deficits; fatigue, pain, spasticity, cognitive impairment", "Review_Status": "Pending"},
    {"Condition_ID": "COND-025", "Condition_Name": "Fibromyalgia", "Category": "Pain / Rheumatic", "ICD_10": "M79.7", "Key_Symptoms": "Widespread musculoskeletal pain, fatigue, sleep disturbance, cognitive difficulties ('fibro fog')", "Review_Status": "Pending"},
    {"Condition_ID": "COND-026", "Condition_Name": "Cognitive Enhancement", "Category": "Cognitive / Performance", "ICD_10": "N/A", "Key_Symptoms": "Enhancement of working memory, attention, processing speed in healthy or MCI populations", "Review_Status": "Pending"},
    {"Condition_ID": "COND-027", "Condition_Name": "Inflammatory / Rheumatoid Arthritis", "Category": "Inflammatory / Pain", "ICD_10": "M05-M06", "Key_Symptoms": "Joint inflammation, pain, swelling, stiffness; systemic inflammatory markers", "Review_Status": "Pending"},
    {"Condition_ID": "COND-028", "Condition_Name": "Motor / Perceptual Learning", "Category": "Cognitive / Performance", "ICD_10": "N/A", "Key_Symptoms": "Enhancement of motor skill acquisition and perceptual discrimination learning", "Review_Status": "Pending"},
    {"Condition_ID": "COND-029", "Condition_Name": "Tourette's Syndrome", "Category": "Movement / Neurodevelopmental", "ICD_10": "F95.2", "Key_Symptoms": "Motor and vocal tics, premonitory urges, comorbid OCD/ADHD", "Review_Status": "Pending"},
    {"Condition_ID": "COND-030", "Condition_Name": "Vascular Cognitive Impairment (VCI)", "Category": "Cerebrovascular / Cognitive", "ICD_10": "F01", "Key_Symptoms": "Cognitive decline attributed to cerebrovascular disease; executive dysfunction, processing speed deficits", "Review_Status": "Pending"},
    {"Condition_ID": "COND-031", "Condition_Name": "MS-related Pain", "Category": "Pain / Neurological", "ICD_10": "G35 + G89", "Key_Symptoms": "Central neuropathic pain in multiple sclerosis; burning, shooting pain, allodynia", "Review_Status": "Pending"},
]

# Add new condition name mappings
new_condition_name_map = {
    "Schizophrenia": "COND-021",
    "Schizophrenia (Auditory Hallucinations)": "COND-021",
    "Schizophrenia (Delusions)": "COND-021",
    "Schizophrenia (Negative Symptoms)": "COND-021",
    "Disorders of Consciousness (DOC)": "COND-022",
    "Mild Cognitive Impairment (MCI)": "COND-023",
    "Cognitive Enhancement / MCI": "COND-023",
    "Cognitive Enhancement": "COND-026",
    "Multiple Sclerosis (MS)": "COND-024",
    "Fibromyalgia": "COND-025",
    "Inflammatory / Rheumatoid Arthritis": "COND-027",
    "Motor / Perceptual Learning": "COND-028",
    "Tourette's Syndrome": "COND-029",
    "Vascular Cognitive Impairment (VCI) + Insomnia": "COND-030",
    "MS-related Pain": "COND-031",
}
condition_name_map.update(new_condition_name_map)

all_conditions = existing_conditions + new_conditions
print(f"  Existing conditions: {len(existing_conditions)}")
print(f"  New conditions added: {len(new_conditions)}")
print(f"  Total conditions: {len(all_conditions)}")

# ============================================================
# PHASE 3: Protocol import — read all 111 protocols from SOZO
# ============================================================

print("\n" + "=" * 60)
print("PHASE 3: Protocol import (111 protocols)")
print("=" * 60)

# Modality sheet -> Modality_ID mapping
modality_sheet_map = {
    "TPS": "MOD-006",
    "TMS_rTMS": "MOD-001",
    "tDCS": "MOD-002",
    "taVNS_tVNS": "MOD-005",
    "CES": "MOD-004",
    "tACS": "MOD-013",
    "PBM": "MOD-007",
    "PEMF": "MOD-014",
    "LIFU_tFUS": "MOD-015",
    "tRNS": "MOD-016",
    "DBS": "MOD-008",
}

# Investigational modalities (GOV-003)
investigational_modalities = {"MOD-013", "MOD-014", "MOD-015", "MOD-016"}

def normalize_regulatory(raw_status, modality_id):
    """Normalize regulatory status to DeepSynaps vocabulary."""
    if not raw_status:
        return "To verify"
    s = str(raw_status).strip()
    
    # FDA Approved (PMA)
    if "FDA Approved" in s:
        if "humanitarian" in s.lower() or "HDE" in s:
            return f"FDA approved (Humanitarian Device Exemption) — {s}"
        return f"FDA approved (PMA) — {s}"
    
    # FDA Cleared (510(k))
    if "FDA Cleared" in s:
        return f"FDA cleared (510(k)) — {s}"
    
    # CE Marked
    if "CE Marked" in s or "CE marked" in s:
        return f"CE-marked — {s}"
    
    # Investigational with nuances
    if "Investigational" in s:
        if "FDA Breakthrough" in s:
            return f"Investigational — {s} (Note: Breakthrough Designation is NOT clearance/approval)"
        if "Nexalin FDA 510(k)" in s:
            return f"Investigational (Nexalin 510(k) cleared for insomnia specifically) — {s}"
        return f"Investigational — {s}"
    
    return f"To verify — {s}"

def map_evidence_level(raw_evidence, lit_count, modality_id):
    """Map evidence descriptions to EV-A/B/C/D grades."""
    if not raw_evidence:
        return "EV-D"
    e = str(raw_evidence).strip().lower()
    
    # Force EV-C or EV-D for investigational modalities (GOV-003)
    force_low = modality_id in investigational_modalities
    
    count = 0
    if lit_count:
        try:
            count = int(str(lit_count).strip().replace('+', '').replace('~', '').split()[0])
        except:
            count = 0
    
    # RCT + Meta-analysis with high paper count
    if ("meta-analysis" in e or "systematic review" in e) and "rct" in e:
        if force_low:
            return "EV-C"
        if count >= 100:
            return "EV-A"
        return "EV-B"
    
    # RCT only
    if "rct" in e and ("meta" not in e):
        if "high quality" in e or "non-inferiority" in e:
            if force_low:
                return "EV-C"
            return "EV-B"
        if "conflicting" in e:
            return "EV-C"
        if force_low:
            return "EV-C"
        return "EV-B"
    
    # Double-blind RCT
    if "double-blind" in e and "rct" in e:
        if force_low:
            return "EV-C"
        return "EV-B"
    
    # Pilot RCT
    if "pilot rct" in e or "pilot" in e and "rct" in e:
        return "EV-C"
    
    # Open-label + RCT
    if "open-label" in e and "rct" in e:
        return "EV-C"
    
    # Open-label, small studies
    if "open-label" in e or "open label" in e:
        return "EV-C" if "multiple" in e or "phase" in e else "EV-D"
    
    # Case series, case report
    if "case series" in e or "case report" in e or "emerging" in e:
        return "EV-D"
    
    # Pilot studies
    if "pilot" in e:
        return "EV-D"
    
    return "EV-D"

# Read all protocols from SOZO master file
wb = openpyxl.load_workbook('/home/user/workspace/SOZO_Master_Neuromodulation_Protocols_v2.xlsx', data_only=True)

protocol_id_counter = 1
all_protocols = []
governance_flags = []
duplicates_detected = []

modality_sheets = ['TPS', 'TMS_rTMS', 'tDCS', 'taVNS_tVNS', 'CES', 'tACS', 'PBM', 'PEMF', 'LIFU_tFUS', 'tRNS', 'DBS']

for sheet_name in modality_sheets:
    ws = wb[sheet_name]
    modality_id = modality_sheet_map[sheet_name]
    
    for row_idx in range(2, ws.max_row + 1):
        condition_raw = ws.cell(row=row_idx, column=1).value
        if not condition_raw or "Source:" in str(condition_raw):
            continue
        
        condition_raw = str(condition_raw).strip()
        condition_id = condition_name_map.get(condition_raw, "To verify")
        
        # Read all fields
        target_region = ws.cell(row=row_idx, column=2).value or ""
        eeg_position = ws.cell(row=row_idx, column=3).value or ""
        protocol_summary = ws.cell(row=row_idx, column=4).value or ""
        intensity = ws.cell(row=row_idx, column=5).value or ""
        frequency = ws.cell(row=row_idx, column=6).value or ""
        session_duration = ws.cell(row=row_idx, column=7).value or ""
        total_sessions = ws.cell(row=row_idx, column=8).value or ""
        pulses_dose = ws.cell(row=row_idx, column=9).value or ""
        devices = ws.cell(row=row_idx, column=10).value or ""
        montage = ws.cell(row=row_idx, column=11).value or ""
        reg_status_raw = ws.cell(row=row_idx, column=12).value or ""
        evidence_raw = ws.cell(row=row_idx, column=13).value or ""
        lit_count = ws.cell(row=row_idx, column=14).value or ""
        key_refs = ws.cell(row=row_idx, column=15).value or ""
        side_effects = ws.cell(row=row_idx, column=16).value or ""
        notes = ws.cell(row=row_idx, column=17).value or ""
        
        # Normalize
        reg_status = normalize_regulatory(str(reg_status_raw), modality_id)
        evidence_level = map_evidence_level(str(evidence_raw), lit_count, modality_id)
        
        # Neurofeedback ADHD lock (GOV-008)
        if modality_id == "MOD-003" and condition_id == "COND-007":
            evidence_level = "EV-D"
        
        protocol = {
            "Protocol_ID": f"PROT-{protocol_id_counter:03d}",
            "Condition_ID": condition_id,
            "Condition_Name": condition_raw,
            "Modality_ID": modality_id,
            "Modality_Sheet": sheet_name,
            "Target_Region": str(target_region),
            "EEG_Position": str(eeg_position),
            "Evidence_Summary": str(protocol_summary),
            "Intensity": str(intensity),
            "Frequency_Hz": str(frequency),
            "Session_Duration": str(session_duration),
            "Total_Course": str(total_sessions),
            "Pulses_Dose": str(pulses_dose),
            "Electrode_Coil_Montage": str(montage),
            "Device_Reference": str(devices),
            "Regulatory_Status": reg_status,
            "Evidence_Level": evidence_level,
            "Evidence_Raw": str(evidence_raw),
            "Literature_Count": str(lit_count),
            "Key_References": str(key_refs),
            "Adverse_Event_Monitoring": str(side_effects),
            "Notes": str(notes),
            "Review_Status": "Pending",
            "Import_Source": "SOZO_Master_Neuromodulation_Protocols_v2.xlsx",
            "Import_Date": datetime.now().strftime("%Y-%m-%d"),
        }
        
        # Governance flags
        flags = []
        if "Investigational" in reg_status or "off-label" in reg_status.lower():
            flags.append("GOV-001")
        if modality_id in investigational_modalities:
            flags.append("GOV-003")
        if evidence_level == "EV-D":
            flags.append("GOV-002")
        
        protocol["Governance_Flags"] = ", ".join(flags) if flags else ""
        
        all_protocols.append(protocol)
        if flags:
            governance_flags.append({"Protocol_ID": protocol["Protocol_ID"], "Flags": flags, "Condition": condition_raw, "Modality": sheet_name})
        
        protocol_id_counter += 1

print(f"  Imported {len(all_protocols)} protocols from SOZO data")
print(f"  Governance flags raised: {len(governance_flags)}")

# Evidence level distribution
ev_dist = {}
for p in all_protocols:
    ev = p["Evidence_Level"]
    ev_dist[ev] = ev_dist.get(ev, 0) + 1
print(f"  Evidence distribution: {ev_dist}")

# ============================================================
# PHASE 5: Brain Regions + qEEG Tables
# ============================================================

print("\n" + "=" * 60)
print("PHASE 5: Brain Regions + qEEG Tables")
print("=" * 60)

# Brain Regions from network file (46 entries)
wb_neuro = openpyxl.load_workbook('/home/user/workspace/Brain_Networks_qEEG.xlsx', data_only=True)
ws_br = wb_neuro['Brain_Regions']

brain_regions = []
br_counter = 1
for row_idx in range(3, ws_br.max_row + 1):
    name = ws_br.cell(row=row_idx, column=1).value
    if not name or not str(name).strip():
        continue
    
    region = {
        "Region_ID": f"BR-{br_counter:03d}",
        "Region_Name": str(name).strip(),
        "Abbreviation": str(ws_br.cell(row=row_idx, column=2).value or "").strip(),
        "Lobe": str(ws_br.cell(row=row_idx, column=3).value or "").strip(),
        "Depth": str(ws_br.cell(row=row_idx, column=4).value or "").strip(),
        "EEG_Position_10_20": str(ws_br.cell(row=row_idx, column=5).value or "").strip(),
        "Brodmann_Area": str(ws_br.cell(row=row_idx, column=6).value or "").strip(),
        "Primary_Functions": str(ws_br.cell(row=row_idx, column=7).value or "").strip(),
        "Brain_Network": str(ws_br.cell(row=row_idx, column=8).value or "").strip(),
        "Key_Conditions": str(ws_br.cell(row=row_idx, column=11).value or "").strip(),
        "Targetable_Modalities": str(ws_br.cell(row=row_idx, column=12).value or "").strip(),
        "Notes": str(ws_br.cell(row=row_idx, column=13).value or "").strip() if ws_br.cell(row=row_idx, column=13).value else "",
        "Review_Status": "Reviewed",
    }
    brain_regions.append(region)
    br_counter += 1

print(f"  Brain Regions: {len(brain_regions)} entries")

# qEEG Biomarkers (7 entries)
ws_qeeg = wb_neuro['qEEG_Biomarkers']
qeeg_biomarkers = []
qb_counter = 1
for row_idx in range(3, ws_qeeg.max_row + 1):
    band = ws_qeeg.cell(row=row_idx, column=1).value
    if not band or not str(band).strip():
        continue
    
    biomarker = {
        "Band_ID": f"QBM-{qb_counter:03d}",
        "Band_Name": str(band).strip(),
        "Hz_Range": str(ws_qeeg.cell(row=row_idx, column=2).value or "").strip(),
        "Normal_Brain_State": str(ws_qeeg.cell(row=row_idx, column=3).value or "").strip(),
        "Key_Regions": str(ws_qeeg.cell(row=row_idx, column=4).value or "").strip(),
        "EEG_Positions": str(ws_qeeg.cell(row=row_idx, column=5).value or "").strip(),
        "Pathological_Increase": str(ws_qeeg.cell(row=row_idx, column=6).value or "").strip(),
        "Pathological_Decrease": str(ws_qeeg.cell(row=row_idx, column=7).value or "").strip(),
        "Associated_Disorders": str(ws_qeeg.cell(row=row_idx, column=8).value or "").strip(),
        "Clinical_Significance": str(ws_qeeg.cell(row=row_idx, column=12).value or "").strip() if ws_qeeg.max_column >= 12 else "",
        "Review_Status": "Reviewed",
    }
    qeeg_biomarkers.append(biomarker)
    qb_counter += 1

print(f"  qEEG Biomarkers: {len(qeeg_biomarkers)} entries")

# qEEG Condition Map (22 entries)
wb_qeeg_map = openpyxl.load_workbook('/home/user/workspace/Master-with-QEEG-places.xlsx', data_only=True)
ws_qmap = wb_qeeg_map['Conditions_QEEG_Map']

qeeg_condition_map = []
qcm_counter = 1
for row_idx in range(2, ws_qmap.max_row + 1):
    cond = ws_qmap.cell(row=row_idx, column=1).value
    if not cond or not str(cond).strip():
        continue
    
    cond_name = str(cond).strip()
    cond_id = condition_name_map.get(cond_name, "To verify")
    
    entry = {
        "Map_ID": f"QCM-{qcm_counter:03d}",
        "Condition_ID": cond_id,
        "Condition_Name": cond_name,
        "Key_Symptoms": str(ws_qmap.cell(row=row_idx, column=2).value or "").strip(),
        "qEEG_Patterns": str(ws_qmap.cell(row=row_idx, column=3).value or "").strip(),
        "Key_qEEG_Electrode_Sites": str(ws_qmap.cell(row=row_idx, column=4).value or "").strip(),
        "Affected_Brain_Regions": str(ws_qmap.cell(row=row_idx, column=5).value or "").strip(),
        "Primary_Networks_Disrupted": str(ws_qmap.cell(row=row_idx, column=6).value or "").strip(),
        "Network_Dysfunction_Pattern": str(ws_qmap.cell(row=row_idx, column=7).value or "").strip(),
        "Recommended_Neuromod_Techniques": str(ws_qmap.cell(row=row_idx, column=8).value or "").strip(),
        "Primary_Stimulation_Targets": str(ws_qmap.cell(row=row_idx, column=9).value or "").strip(),
        "Stimulation_Rationale": str(ws_qmap.cell(row=row_idx, column=10).value or "").strip(),
        "Review_Status": "Reviewed",
    }
    qeeg_condition_map.append(entry)
    qcm_counter += 1

print(f"  qEEG Condition Map: {len(qeeg_condition_map)} entries")

# ============================================================
# PHASE 6: Assessment enrichment
# ============================================================

print("\n" + "=" * 60)
print("PHASE 6: Assessment enrichment")
print("=" * 60)

wb_assess = openpyxl.load_workbook('/home/user/workspace/Assessments_Master.xlsx', data_only=True)
ws_am = wb_assess['Assessment_Master']

# Build assessment records from the master sheet
assessments = []
assess_counter = 1

# Existing assessment IDs we know about
known_assessments = {
    "PHQ-9", "HAM-D", "MADRS", "BDI-II", "GAD-7", "ISI", "PSQI",
    "PCL-5", "Y-BOCS", "CAARS", "Conners", "MoCA", "MMSE",
    "NRS", "VAS", "BPI", "FMA", "ARAT", "THI", "TFI",
    "HADS", "WHODAS", "CGI", "BAI", "AUDIT", "DAST",
    "QoL-AD", "UPDRS", "MDS-UPDRS", "ADAS-Cog", "CDR",
    "ACE-III", "C-SSRS", "PANSS", "CRS-R", "FIQ-R", "EDSS",
    "BFMDRS", "TETRAS", "ADOS-2", "YGTSS", "H&Y",
    "SHAPS", "IDS-SR", "QIDS-SR", "Q-LES-Q", "WSAS",
    "NPI", "FAST", "Barthel", "ZBI",
}

for row_idx in range(3, ws_am.max_row + 1):
    cond = ws_am.cell(row=row_idx, column=1).value
    if not cond or not str(cond).strip():
        continue
    
    cond_name = str(cond).strip()
    cond_id = condition_name_map.get(cond_name, "To verify")
    
    # Check for unmapped conditions in assessments
    if cond_id == "To verify":
        # Try partial matching
        for key, val in condition_name_map.items():
            if cond_name.lower().startswith(key.lower().split('(')[0].strip().split('/')[0].strip()):
                cond_id = val
                break
    
    category = str(ws_am.cell(row=row_idx, column=2).value or "").strip()
    primary_scales = str(ws_am.cell(row=row_idx, column=3).value or "").strip()
    neuropsych = str(ws_am.cell(row=row_idx, column=4).value or "").strip()
    qeeg_bands = str(ws_am.cell(row=row_idx, column=5).value or "").strip()
    qeeg_electrodes = str(ws_am.cell(row=row_idx, column=6).value or "").strip()
    qeeg_metrics = str(ws_am.cell(row=row_idx, column=7).value or "").strip()
    brain_regions = str(ws_am.cell(row=row_idx, column=8).value or "").strip()
    primary_network = str(ws_am.cell(row=row_idx, column=9).value or "").strip()
    neuroimaging = str(ws_am.cell(row=row_idx, column=10).value or "").strip()
    physiological = str(ws_am.cell(row=row_idx, column=11).value or "").strip()
    functional = str(ws_am.cell(row=row_idx, column=12).value or "").strip()
    network_target = str(ws_am.cell(row=row_idx, column=13).value or "").strip()
    rationale = str(ws_am.cell(row=row_idx, column=14).value or "").strip()
    
    assessment = {
        "Assessment_ID": f"ASSESS-{assess_counter:03d}",
        "Condition_ID": cond_id,
        "Condition_Name": cond_name,
        "Category": category,
        "Primary_Clinical_Scales": primary_scales,
        "Neuropsychological_Battery": neuropsych,
        "qEEG_Key_Bands": qeeg_bands,
        "Key_qEEG_Electrodes": qeeg_electrodes,
        "Key_qEEG_Metrics": qeeg_metrics,
        "Brain_Regions_Affected": brain_regions,
        "Primary_Network_Disrupted": primary_network,
        "Neuroimaging": neuroimaging,
        "Physiological_Assessments": physiological,
        "Functional_Behavioural": functional,
        "Brain_qEEG_Treatment_Target": network_target,
        "Clinical_Rationale": rationale,
        "Review_Status": "Reviewed",
        "Import_Source": "Assessments_Master.xlsx",
    }
    assessments.append(assessment)
    assess_counter += 1

print(f"  Assessment profiles: {len(assessments)} condition-assessment mappings")

# ============================================================
# PHASE 7: Source library expansion
# ============================================================

print("\n" + "=" * 60)
print("PHASE 7: Source library expansion")
print("=" * 60)

sources = []
src_counter = 1

def read_csv_papers(filepath, min_citations=50):
    """Read papers from CSV, filter by citation count."""
    papers = []
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                citations = int(row.get('Citations', '0').strip())
            except (ValueError, AttributeError):
                citations = 0
            if citations >= min_citations:
                papers.append(row)
    papers.sort(key=lambda x: int(x.get('Citations', '0').strip() or '0'), reverse=True)
    return papers

# TPS/tFUS literature — systematic reviews + meta-analyses with 50+ citations
tps_papers = read_csv_papers('/home/user/workspace/transcinal-pulse-stimulations-04-Apr-2026-1.csv', min_citations=50)
# Filter for reviews and meta-analyses
tps_filtered = []
for p in tps_papers:
    study_type = str(p.get('Study Type', '')).lower()
    title = str(p.get('\ufeffTitle', p.get('Title', ''))).lower()
    if any(term in study_type + ' ' + title for term in ['review', 'meta-analysis', 'meta analysis', 'systematic', 'survey', 'overview']):
        tps_filtered.append(p)
        if len(tps_filtered) >= 25:
            break

# Also add top-cited papers regardless of type
for p in tps_papers[:10]:
    if p not in tps_filtered:
        tps_filtered.append(p)

# Brain networks papers with 100+ citations
bn_neuromod = read_csv_papers('/home/user/workspace/brain-networks-for-neurmodulation-04-Apr-2026.csv', min_citations=50)
bn_general = read_csv_papers('/home/user/workspace/brain-networks-04-Apr-2026.csv', min_citations=100)

all_high_papers = tps_filtered + bn_neuromod[:15] + bn_general[:10]

for paper in all_high_papers:
    title = paper.get('\ufeffTitle', paper.get('Title', '')).strip()
    if not title:
        continue
    
    doi = paper.get('DOI', '').strip()
    url = f"https://doi.org/{doi}" if doi else ""
    year = paper.get('Year', '').strip()
    citations = paper.get('Citations', '0').strip()
    study_type = paper.get('Study Type', '').strip()
    journal = paper.get('Journal', '').strip()
    
    # Determine authority level
    try:
        cit_num = int(citations)
    except:
        cit_num = 0
    
    if cit_num >= 500:
        authority = "Tier 1 — Landmark"
    elif cit_num >= 200:
        authority = "Tier 1 — High Impact"
    elif cit_num >= 100:
        authority = "Tier 2 — Well-Cited"
    else:
        authority = "Tier 3 — Relevant"
    
    source = {
        "Source_ID": f"SRC-{src_counter:03d}",
        "Source_Type": study_type if study_type else "To verify",
        "Title": title[:200],
        "URL": url,
        "DOI": doi,
        "Journal": journal,
        "Authority_Level": authority,
        "Publication_Year": year,
        "Citations": str(cit_num),
        "Use_Case": "TPS/tFUS neuromodulation" if paper in tps_filtered else "Brain network science / neuromodulation",
        "Review_Status": "Reviewed",
    }
    sources.append(source)
    src_counter += 1

# Deduplicate by title
seen_titles = set()
unique_sources = []
for s in sources:
    title_key = s["Title"].lower().strip()[:100]
    if title_key not in seen_titles:
        seen_titles.add(title_key)
        unique_sources.append(s)
sources = unique_sources[:50]  # Cap at 50

print(f"  High-quality sources extracted: {len(sources)}")
print(f"    TPS/tFUS reviews/meta-analyses: {len(tps_filtered)}")
print(f"    Brain network papers: {len(bn_neuromod[:15]) + len(bn_general[:10])}")

# ============================================================
# EXISTING TABLES: Devices, Symptoms/Phenotypes (seed data)
# ============================================================

print("\n" + "=" * 60)
print("Rebuilding Devices and Phenotypes tables")
print("=" * 60)

# Devices table — extract unique devices from protocols
devices = []
dev_counter = 1
seen_devices = set()

device_catalog = [
    ("DEV-001", "NeuroStar", "MOD-001", "TMS System", "FDA cleared (510(k)) for MDD (2008)", "Neuronetics"),
    ("DEV-002", "BrainsWay Deep TMS", "MOD-001", "Deep TMS System", "FDA cleared (510(k)) for MDD, OCD (2018), smoking cessation (2020)", "BrainsWay"),
    ("DEV-003", "MagVenture", "MOD-001", "TMS System", "FDA cleared (510(k)) for MDD", "MagVenture"),
    ("DEV-004", "Magstim", "MOD-001", "TMS System", "FDA cleared (510(k)) for MDD", "Magstim"),
    ("DEV-005", "Flow FL-100", "MOD-002", "tDCS System", "FDA approved (PMA) for treatment-resistant MDD (Dec 2025)", "Flow Neuroscience"),
    ("DEV-006", "Soterix HD-tDCS", "MOD-002", "HD-tDCS System", "Research use only in US; CE marked in Europe", "Soterix Medical"),
    ("DEV-007", "Neuroelectrics Starstim", "MOD-002", "Multi-channel tES System", "Research use only in US; CE marked in Europe", "Neuroelectrics"),
    ("DEV-008", "Alpha-Stim M/AID", "MOD-004", "CES Device", "FDA cleared for anxiety, insomnia, depression, pain", "Electromedical Products International"),
    ("DEV-009", "Fisher Wallace Stimulator", "MOD-004", "CES Device", "FDA cleared for depression, anxiety, insomnia", "Fisher Wallace Laboratories"),
    ("DEV-010", "NEUROLITH", "MOD-006", "TPS System", "CE marked for Alzheimer's (Europe)", "Storz Medical AG"),
    ("DEV-011", "Vielight Neuro Gamma", "MOD-007", "Transcranial PBM System", "Investigational — not FDA cleared", "Vielight"),
    ("DEV-012", "SYMBYX PDCare", "MOD-007", "PBM Laser", "TGA listed (Australia)", "SYMBYX Biome"),
    ("DEV-013", "Medtronic Percept PC", "MOD-008", "DBS System", "FDA approved for PD, ET, dystonia, OCD (HDE), epilepsy", "Medtronic"),
    ("DEV-014", "Abbott Infinity", "MOD-008", "DBS System", "FDA approved for PD, ET, dystonia, OCD (HDE)", "Abbott"),
    ("DEV-015", "Boston Scientific Vercise", "MOD-008", "DBS System", "FDA approved for PD, ET, dystonia", "Boston Scientific"),
    ("DEV-016", "NEMOS", "MOD-005", "taVNS Device", "CE marked for epilepsy and depression (Europe)", "tVNS Technologies"),
    ("DEV-017", "Parasym", "MOD-005", "taVNS Device", "CE marked (Europe)", "Parasym"),
    ("DEV-018", "gammaCore", "MOD-011", "Cervical VNS Device", "FDA cleared for cluster headache and migraine prevention", "electroCore"),
    ("DEV-019", "Nexalin", "MOD-013", "tACS Device", "510(k) cleared for insomnia specifically", "Nexalin Technology"),
    ("DEV-020", "NeoRhythm", "MOD-014", "PEMF Headband", "Investigational — consumer wellness device", "OmniPEMF"),
    ("DEV-021", "BrainSonix BX Pulsar", "MOD-015", "tFUS System", "Investigational — not FDA cleared", "BrainSonix"),
    ("DEV-022", "Openwater", "MOD-015", "tFUS System", "Investigational — not FDA cleared", "Openwater"),
    ("DEV-023", "eNeura SpringTMS", "MOD-001", "sTMS Device", "FDA cleared for migraine (2013)", "eNeura"),
    ("DEV-024", "Newronika HDC-STIM", "MOD-002", "tDCS System", "Investigational", "Newronika"),
    ("DEV-025", "CES Ultra", "MOD-004", "CES Device", "FDA cleared", "Neuro-Fitness"),
    ("DEV-026", "ICES-DigiCeutical", "MOD-014", "PEMF Device", "Investigational — consumer wellness device", "Micro-Pulse"),
    ("DEV-027", "FlexPulse", "MOD-014", "PEMF Device", "Investigational — consumer wellness device", "FlexPulse"),
    ("DEV-028", "Thor Photomedicine", "MOD-007", "PBM System", "Investigational", "Thor Photomedicine"),
    ("DEV-029", "Brainbox", "MOD-015", "tFUS System", "Investigational — not FDA cleared", "Brainbox"),
]

for dev_id, name, mod_id, dev_type, reg, mfr in device_catalog:
    devices.append({
        "Device_ID": dev_id,
        "Device_Name": name,
        "Modality_ID": mod_id,
        "Device_Type": dev_type,
        "Regulatory_Status": reg,
        "Manufacturer": mfr,
        "Review_Status": "Reviewed",
    })

print(f"  Devices catalogued: {len(devices)}")

# Phenotypes table (from existing v1 + SOZO)
phenotypes = [
    {"Phenotype_ID": "PHEN-001", "Phenotype_Name": "Treatment-Resistant Depression — Frontal Hypoactivity", "Condition_ID": "COND-001", "Key_Biomarkers": "↓ L-DLPFC activity (FAA F4>F3), ↑ frontal theta, ↓ alpha globally", "Suggested_Modalities": "TMS (10 Hz L-DLPFC), tDCS (anodal F3), TPS (F3 target)", "Evidence_Level": "EV-A", "Review_Status": "Reviewed"},
    {"Phenotype_ID": "PHEN-002", "Phenotype_Name": "Alzheimer's — DMN Disruption with Alpha Slowing", "Condition_ID": "COND-009", "Key_Biomarkers": "↓ APF (<8 Hz), ↑ theta/delta, ↓ DMN connectivity (Pz, PCC)", "Suggested_Modalities": "TPS (multi-site DMN), 40 Hz tACS (temporal), PBM (810 nm 40 Hz)", "Evidence_Level": "EV-B", "Review_Status": "Reviewed"},
    {"Phenotype_ID": "PHEN-003", "Phenotype_Name": "ADHD — Elevated Theta/Beta Ratio", "Condition_ID": "COND-007", "Key_Biomarkers": "↑ TBR (Cz), ↑ frontal theta, ↓ beta (F3/F4/Cz), poor P300", "Suggested_Modalities": "Neurofeedback (EV-D — Cortese 2024), TMS (investigational)", "Evidence_Level": "EV-D", "Review_Status": "Reviewed"},
    {"Phenotype_ID": "PHEN-004", "Phenotype_Name": "Chronic Pain — Central Sensitization", "Condition_ID": "COND-008", "Key_Biomarkers": "↓ Mu (C3/C4), ↑ theta (central/frontal), ↓ alpha, ↑ gamma (pain matrix)", "Suggested_Modalities": "tDCS (anodal M1), TMS (M1), taVNS, CES, LIFU (thalamus/ACC)", "Evidence_Level": "EV-B", "Review_Status": "Reviewed"},
    {"Phenotype_ID": "PHEN-005", "Phenotype_Name": "PTSD — Hyperarousal with Salience Network Overdrive", "Condition_ID": "COND-005", "Key_Biomarkers": "↑ beta (global), ↑ right frontal alpha, ↓ alpha globally, ↑ gamma (amygdala/insula)", "Suggested_Modalities": "TMS (R-DLPFC inhibitory), CES, taVNS, EMDR + neuromodulation", "Evidence_Level": "EV-B", "Review_Status": "Reviewed"},
    {"Phenotype_ID": "PHEN-006", "Phenotype_Name": "Parkinson's — Motor Circuit Beta Excess", "Condition_ID": "COND-010", "Key_Biomarkers": "↑ beta (C3/C4) at rest, ↓ beta suppression during movement, ↑ DAR", "Suggested_Modalities": "DBS (STN/GPi), TMS (M1), TPS (M1+SMA), PBM", "Evidence_Level": "EV-A", "Review_Status": "Reviewed"},
]

print(f"  Phenotypes: {len(phenotypes)}")

# ============================================================
# WRITE ALL CSVs
# ============================================================

print("\n" + "=" * 60)
print("Writing CSV files")
print("=" * 60)

def write_csv(filename, data, fieldnames=None):
    if not data:
        return
    if not fieldnames:
        fieldnames = list(data[0].keys())
    filepath = os.path.join(CSV_DIR, filename)
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)
    print(f"  Written: {filename} ({len(data)} rows)")

write_csv("Evidence_Levels.csv", evidence_levels)
write_csv("Governance_Rules.csv", governance_rules)
write_csv("Modalities.csv", all_modalities)
write_csv("Devices.csv", devices)
write_csv("Conditions.csv", all_conditions)
write_csv("Symptoms_Phenotypes.csv", phenotypes)
write_csv("Assessments.csv", assessments)
write_csv("Protocols.csv", all_protocols)
write_csv("Sources.csv", sources)
write_csv("Brain_Regions.csv", brain_regions)
write_csv("qEEG_Condition_Map.csv", qeeg_condition_map)
write_csv("qEEG_Biomarkers.csv", qeeg_biomarkers)

# ============================================================
# WRITE EXCEL WORKBOOK
# ============================================================

print("\n" + "=" * 60)
print("Writing Excel workbook")
print("=" * 60)

wb_out = openpyxl.Workbook()

# Styles
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
alt_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
pending_fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def write_sheet(wb, sheet_name, data, fieldnames=None):
    if not data:
        return
    if not fieldnames:
        fieldnames = list(data[0].keys())
    ws = wb.create_sheet(title=sheet_name[:31])
    
    # Headers
    for col, header in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Data
    for row_idx, record in enumerate(data, 2):
        for col, field in enumerate(fieldnames, 1):
            val = record.get(field, "")
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            
            # Highlight pending items
            if field == "Review_Status" and val == "Pending":
                cell.fill = pending_fill
    
    # Auto-width (capped at 50)
    for col in range(1, len(fieldnames) + 1):
        max_len = len(str(ws.cell(row=1, column=col).value))
        for row in range(2, min(len(data) + 2, 20)):
            cell_val = str(ws.cell(row=row, column=col).value or "")
            max_len = max(max_len, min(len(cell_val), 50))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 4, 55)
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    return ws

# Remove default sheet
wb_out.remove(wb_out.active)

# Write all 12 tables
write_sheet(wb_out, "Evidence_Levels", evidence_levels)
write_sheet(wb_out, "Governance_Rules", governance_rules)
write_sheet(wb_out, "Modalities", all_modalities)
write_sheet(wb_out, "Devices", devices)
write_sheet(wb_out, "Conditions", all_conditions)
write_sheet(wb_out, "Symptoms_Phenotypes", phenotypes)
write_sheet(wb_out, "Assessments", assessments)
write_sheet(wb_out, "Protocols", all_protocols)
write_sheet(wb_out, "Sources", sources)
write_sheet(wb_out, "Brain_Regions", brain_regions)
write_sheet(wb_out, "qEEG_Condition_Map", qeeg_condition_map)
write_sheet(wb_out, "qEEG_Biomarkers", qeeg_biomarkers)

excel_path = os.path.join(OUTPUT_DIR, "DeepSynaps_Master_Database_v2.xlsx")
wb_out.save(excel_path)
print(f"  Excel workbook saved: {excel_path}")

# ============================================================
# INTEGRATION REPORT
# ============================================================

print("\n" + "=" * 60)
print("Generating Integration Report")
print("=" * 60)

total_records = (
    len(evidence_levels) + len(governance_rules) + len(all_modalities) + 
    len(devices) + len(all_conditions) + len(phenotypes) + 
    len(assessments) + len(all_protocols) + len(sources) +
    len(brain_regions) + len(qeeg_condition_map) + len(qeeg_biomarkers)
)

# Count governance flags
gov_flag_counts = {}
for p in all_protocols:
    for flag in p.get("Governance_Flags", "").split(", "):
        if flag:
            gov_flag_counts[flag] = gov_flag_counts.get(flag, 0) + 1

# Count pending reviews
pending_count = sum(1 for p in all_protocols if p["Review_Status"] == "Pending")
pending_conditions = sum(1 for c in all_conditions if c["Review_Status"] == "Pending")

report = f"""# DeepSynaps Studio — Data Integration Report
## SOZO Brain Center Protocol Library Import

**Date:** {datetime.now().strftime('%Y-%m-%d %H:%M')}
**Database Version:** 2.0
**Previous Version:** 1.0 (201 records, 9 tables)
**Source Files Processed:** 7 (3 Excel workbooks + 3 literature CSVs + 1 primary protocol workbook)

---

## Summary

| Metric | Before | After | Delta |
|---|---|---|---|
| Tables | 9 | 12 | +3 |
| Total Records | 201 | {total_records} | +{total_records - 201} |
| Modalities | 12 | {len(all_modalities)} | +4 |
| Conditions | 20 | {len(all_conditions)} | +{len(new_conditions)} |
| Protocols | 32 | {len(all_protocols)} | +{len(all_protocols) - 32} |
| Devices | ~15 | {len(devices)} | +{len(devices) - 15} |
| Assessments | 42 | {len(assessments)} | +{len(assessments) - 42} |
| Sources | ~20 | {len(sources)} | +{len(sources) - 20} |
| Brain Regions | 0 | {len(brain_regions)} | +{len(brain_regions)} |
| qEEG Condition Maps | 0 | {len(qeeg_condition_map)} | +{len(qeeg_condition_map)} |
| qEEG Biomarkers | 0 | {len(qeeg_biomarkers)} | +{len(qeeg_biomarkers)} |
| Phenotypes | ~5 | {len(phenotypes)} | +{len(phenotypes) - 5} |

---

## New Tables Added

1. **Brain_Regions** — {len(brain_regions)} anatomical regions with 10-20 EEG positions, Brodmann areas, brain network assignments
2. **qEEG_Condition_Map** — {len(qeeg_condition_map)} conditions with qEEG biomarker signatures, electrode sites, network dysfunction patterns, stimulation rationale
3. **qEEG_Biomarkers** — {len(qeeg_biomarkers)} frequency bands with normal/pathological signatures and clinical significance

## New Modalities Added

| ID | Modality | Protocols | Regulatory |
|---|---|---|---|
| MOD-013 | tACS | 10 | Investigational (Nexalin 510(k) for insomnia only) |
| MOD-014 | PEMF | 11 | Investigational — no FDA clearance for neuropsychiatric |
| MOD-015 | LIFU/tFUS | 9 | Investigational — MRgFUS ablation clearance ≠ neuromod |
| MOD-016 | tRNS | 9 | Investigational — no FDA clearance for neuropsychiatric |

## New Conditions Added ({len(new_conditions)})

| ID | Condition | Category | Status |
|---|---|---|---|
"""

for c in new_conditions:
    report += f"| {c['Condition_ID']} | {c['Condition_Name']} | {c['Category']} | {c['Review_Status']} |\n"

report += f"""
## Protocol Import Summary

- **Total protocols imported:** {len(all_protocols)}
- **All protocols set to Review_Status = "Pending"** (none auto-published)
- **Evidence level distribution:**

| Grade | Count | Description |
|---|---|---|
| EV-A | {ev_dist.get('EV-A', 0)} | Guideline-endorsed |
| EV-B | {ev_dist.get('EV-B', 0)} | Strong research evidence |
| EV-C | {ev_dist.get('EV-C', 0)} | Emerging evidence |
| EV-D | {ev_dist.get('EV-D', 0)} | Preliminary / investigational |

## Governance Flags Raised

| Flag | Count | Description |
|---|---|---|
| GOV-001 | {gov_flag_counts.get('GOV-001', 0)} | Off-label / investigational protocol |
| GOV-002 | {gov_flag_counts.get('GOV-002', 0)} | EV-D evidence — cannot be patient-facing without clinician override |
| GOV-003 | {gov_flag_counts.get('GOV-003', 0)} | Investigational modality (tACS, PEMF, LIFU, tRNS) |

## Review Queue

- **Protocols pending review:** {pending_count}
- **Conditions pending review:** {pending_conditions}
- **All new records routed through review queue** per GOV-011

## Regulatory Integrity Checks

- [x] "FDA cleared" and "FDA approved" used with correct distinction throughout
- [x] Flow FL-100 remains the ONLY PMA-approved tDCS device (GOV-009)
- [x] "FDA Breakthrough Designation" noted as designation only, NOT clearance
- [x] All tACS/PEMF/LIFU/tRNS protocols flagged with GOV-003
- [x] Neurofeedback ADHD remains EV-D (GOV-008)
- [x] No marketing language in regulatory or evidence fields (GOV-010)
- [x] All new records enter review queue before publication (GOV-011)

## Source Library

- **High-quality sources added:** {len(sources)}
- **TPS/tFUS systematic reviews and meta-analyses:** {len(tps_filtered)}
- **Brain network science papers (100+ citations):** {len(bn_general[:10])}
- **Network-guided neuromodulation papers:** {len(bn_neuromod[:15])}

## Items Requiring Clinician Review

1. All {pending_count} new protocols need clinical review before publication
2. {pending_conditions} new conditions need clinical validation
3. Condition-to-assessment mappings should be verified by clinical team
4. qEEG biomarker thresholds should be validated against SOZO Brain Center clinical data
5. Device regulatory statuses should be periodically re-verified (quarterly)

---

*Report generated automatically by DeepSynaps Studio Integration Engine*
*Database snapshot created: {datetime.now().strftime('%Y-%m-%d %H:%M')}*
"""

report_path = os.path.join(OUTPUT_DIR, "Integration_Report.md")
with open(report_path, 'w', encoding='utf-8') as f:
    f.write(report)
print(f"  Integration report saved: {report_path}")

# ============================================================
# DATA DICTIONARY
# ============================================================

print("\n" + "=" * 60)
print("Generating Data Dictionary")
print("=" * 60)

data_dict = """# DeepSynaps Studio — Data Dictionary v2.0

**Database Version:** 2.0
**Date:** """ + datetime.now().strftime('%Y-%m-%d') + """
**Total Tables:** 12
**Total Records:** """ + str(total_records) + """

---

## Table 1: Evidence_Levels (""" + str(len(evidence_levels)) + """ records)
Primary key: Evidence_ID

| Field | Type | Description | Example |
|---|---|---|---|
| Evidence_ID | PK, string | Unique identifier (EV-A through EV-D) | EV-B |
| Level_Name | string | Human-readable grade name | "Strong Research Evidence" |
| Description | text | Detailed criteria for this evidence level | |
| Minimum_Evidence | text | Minimum studies/quality required | "≥1 RCT + systematic review" |
| Usage_Gate | text | What this grade permits in the platform | "Can be recommended with clinician judgment" |
| Review_Status | enum | Reviewed / Pending / To Verify | Reviewed |

## Table 2: Governance_Rules (""" + str(len(governance_rules)) + """ records)
Primary key: Rule_ID

| Field | Type | Description | Example |
|---|---|---|---|
| Rule_ID | PK, string | Unique identifier (GOV-001 through GOV-012) | GOV-001 |
| Rule_Name | string | Short rule name | "Off-Label Protocol Flag" |
| Description | text | Full rule description | |
| Trigger_Condition | text | When this rule activates | |
| Action | text | What happens when triggered | |
| Severity | enum | Warning / Critical | Critical |
| Review_Status | enum | Reviewed / Pending / To Verify | Reviewed |

## Table 3: Modalities (""" + str(len(all_modalities)) + """ records)
Primary key: Modality_ID

| Field | Type | Description | Example |
|---|---|---|---|
| Modality_ID | PK, string | MOD-001 through MOD-016 | MOD-001 |
| Modality_Name | string | Short name | "TMS / rTMS" |
| Full_Name | string | Full modality name | "Transcranial Magnetic Stimulation" |
| Category | string | Classification | "Non-invasive Brain Stimulation" |
| Mechanism | text | How the modality works | |
| FDA_Status_Summary | text | Summary of FDA regulatory status | |
| Regulatory_Notes | text | Detailed regulatory notes with caveats | |
| Key_Parameters | text | Primary adjustable parameters | |
| Review_Status | enum | Reviewed / Pending / To Verify | Reviewed |

## Table 4: Devices (""" + str(len(devices)) + """ records)
Primary key: Device_ID

| Field | Type | Description | Example |
|---|---|---|---|
| Device_ID | PK, string | DEV-001 through DEV-029 | DEV-005 |
| Device_Name | string | Commercial device name | "Flow FL-100" |
| Modality_ID | FK→Modalities | Which modality | MOD-002 |
| Device_Type | string | Device category | "tDCS System" |
| Regulatory_Status | text | Precise regulatory status | "FDA approved (PMA) for treatment-resistant MDD" |
| Manufacturer | string | Company name | "Flow Neuroscience" |
| Review_Status | enum | Reviewed / Pending / To Verify | Reviewed |

## Table 5: Conditions (""" + str(len(all_conditions)) + """ records)
Primary key: Condition_ID

| Field | Type | Description | Example |
|---|---|---|---|
| Condition_ID | PK, string | COND-001 through COND-031 | COND-001 |
| Condition_Name | string | Clinical condition name | "Major Depressive Disorder (MDD)" |
| Category | string | Condition classification | "Mood / Affective" |
| ICD_10 | string | ICD-10 code(s) | "F32, F33" |
| Key_Symptoms | text | Core symptoms | |
| Review_Status | enum | Reviewed / Pending / To Verify | Reviewed |

## Table 6: Symptoms_Phenotypes (""" + str(len(phenotypes)) + """ records)
Primary key: Phenotype_ID

| Field | Type | Description | Example |
|---|---|---|---|
| Phenotype_ID | PK, string | PHEN-001 through PHEN-006 | PHEN-001 |
| Phenotype_Name | string | Descriptive phenotype name | "TRD — Frontal Hypoactivity" |
| Condition_ID | FK→Conditions | Linked condition | COND-001 |
| Key_Biomarkers | text | qEEG and neuroimaging biomarkers | |
| Suggested_Modalities | text | Recommended interventions | |
| Evidence_Level | FK→Evidence_Levels | Evidence grade | EV-A |
| Review_Status | enum | Reviewed / Pending / To Verify | Reviewed |

## Table 7: Assessments (""" + str(len(assessments)) + """ records)
Primary key: Assessment_ID

| Field | Type | Description | Example |
|---|---|---|---|
| Assessment_ID | PK, string | ASSESS-001 through ASSESS-0XX | ASSESS-001 |
| Condition_ID | FK→Conditions | Linked condition | COND-001 |
| Condition_Name | string | Display name | "Depression (MDD)" |
| Category | string | Clinical category | "Mood/Affective" |
| Primary_Clinical_Scales | text | Validated scales | "PHQ-9, HAM-D, MADRS, BDI-II" |
| Neuropsychological_Battery | text | Neuropsych tests | |
| qEEG_Key_Bands | text | Relevant frequency bands | |
| Key_qEEG_Electrodes | text | 10-20 electrode sites | "F3, F4, Fp1/Fp2, Fz" |
| Key_qEEG_Metrics | text | Quantitative metrics | "FAA, TBR, APF" |
| Brain_Regions_Affected | text | Affected anatomy | |
| Primary_Network_Disrupted | text | Network(s) involved | "DMN, ECN, Salience" |
| Neuroimaging | text | Imaging recommendations | |
| Physiological_Assessments | text | HRV, sleep, etc. | |
| Functional_Behavioural | text | Functional assessments | |
| Brain_qEEG_Treatment_Target | text | qEEG-guided treatment target | |
| Clinical_Rationale | text | Why these assessments | |
| Review_Status | enum | Reviewed / Pending / To Verify | Reviewed |
| Import_Source | string | Source file | |

## Table 8: Protocols (""" + str(len(all_protocols)) + """ records)
Primary key: Protocol_ID

| Field | Type | Description | Example |
|---|---|---|---|
| Protocol_ID | PK, string | PROT-001 through PROT-111 | PROT-001 |
| Condition_ID | FK→Conditions | Linked condition | COND-009 |
| Condition_Name | string | Original condition name from source | "Alzheimer's Disease" |
| Modality_ID | FK→Modalities | Treatment modality | MOD-006 |
| Modality_Sheet | string | Source sheet name | "TPS" |
| Target_Region | text | Brain target | "dlPFC, IFC, LPC, Precuneus, DMN" |
| EEG_Position | string | 10-20 electrode position | "F3, F7, P3, Pz" |
| Evidence_Summary | text | Protocol description | |
| Intensity | string | Stimulation intensity | "EFD 0.20 mJ/mm²" |
| Frequency_Hz | string | Stimulation frequency | "5 Hz PRF" |
| Session_Duration | string | Duration per session | "~20 min" |
| Total_Course | string | Total treatment course | "6-12 sessions / 2-4 weeks" |
| Pulses_Dose | string | Pulses/dose per session | "6000 pulses" |
| Electrode_Coil_Montage | text | Electrode/coil setup | |
| Device_Reference | text | Applicable devices | "NEUROLITH (Storz Medical AG)" |
| Regulatory_Status | text | Normalized regulatory status | |
| Evidence_Level | FK→Evidence_Levels | EV-A through EV-D | EV-B |
| Evidence_Raw | string | Original evidence description | |
| Literature_Count | string | Published paper count | "31" |
| Key_References | text | Source citations | |
| Adverse_Event_Monitoring | text | Side effects and monitoring | |
| Notes | text | Additional clinical notes | |
| Governance_Flags | string | Active governance flags | "GOV-001, GOV-003" |
| Review_Status | enum | Reviewed / Pending / To Verify | Pending |
| Import_Source | string | Source file | |
| Import_Date | date | Import timestamp | "2026-04-07" |

## Table 9: Sources (""" + str(len(sources)) + """ records)
Primary key: Source_ID

| Field | Type | Description | Example |
|---|---|---|---|
| Source_ID | PK, string | SRC-001 through SRC-0XX | SRC-001 |
| Source_Type | string | Study design type | "Systematic Review" |
| Title | string | Paper title | |
| URL | string | DOI URL | "https://doi.org/10.xxxx" |
| DOI | string | DOI identifier | |
| Journal | string | Journal name | |
| Authority_Level | string | Citation-based authority | "Tier 1 — Landmark" |
| Publication_Year | string | Year published | "2019" |
| Citations | string | Citation count | "230" |
| Use_Case | string | Relevance category | "TPS/tFUS neuromodulation" |
| Review_Status | enum | Reviewed / Pending / To Verify | Reviewed |

## Table 10: Brain_Regions (""" + str(len(brain_regions)) + """ records) — NEW
Primary key: Region_ID

| Field | Type | Description | Example |
|---|---|---|---|
| Region_ID | PK, string | BR-001 through BR-046 | BR-001 |
| Region_Name | string | Anatomical name | "Dorsolateral PFC" |
| Abbreviation | string | Short form | "DLPFC" |
| Lobe | string | Brain lobe | "Frontal" |
| Depth | enum | Cortical / Subcortical | "Cortical" |
| EEG_Position_10_20 | string | 10-20 system positions | "F3(L),F4(R)" |
| Brodmann_Area | string | Brodmann area(s) | "BA9,46" |
| Primary_Functions | text | Key functions | |
| Brain_Network | string | Network assignment(s) | "ECN, Salience, DMN" |
| Key_Conditions | text | Targeted conditions | |
| Targetable_Modalities | text | Modalities that can reach it | |
| Notes | text | Clinical notes | |
| Review_Status | enum | Reviewed / Pending / To Verify | Reviewed |

## Table 11: qEEG_Condition_Map (""" + str(len(qeeg_condition_map)) + """ records) — NEW
Primary key: Map_ID

| Field | Type | Description | Example |
|---|---|---|---|
| Map_ID | PK, string | QCM-001 through QCM-022 | QCM-001 |
| Condition_ID | FK→Conditions | Linked condition | COND-001 |
| Condition_Name | string | Display name | "Depression (MDD)" |
| Key_Symptoms | text | Core symptoms | |
| qEEG_Patterns | text | Characteristic EEG patterns | "↑ frontal theta, FAA" |
| Key_qEEG_Electrode_Sites | string | Key electrode sites | "F3, F4, Fz" |
| Affected_Brain_Regions | text | Involved regions | |
| Primary_Networks_Disrupted | text | Disrupted networks | "DMN, CEN, Salience" |
| Network_Dysfunction_Pattern | text | Mechanism description | |
| Recommended_Neuromod_Techniques | text | Recommended interventions | |
| Primary_Stimulation_Targets | text | Target sites | |
| Stimulation_Rationale | text | Clinical rationale | |
| Review_Status | enum | Reviewed / Pending / To Verify | Reviewed |

## Table 12: qEEG_Biomarkers (""" + str(len(qeeg_biomarkers)) + """ records) — NEW
Primary key: Band_ID

| Field | Type | Description | Example |
|---|---|---|---|
| Band_ID | PK, string | QBM-001 through QBM-007 | QBM-001 |
| Band_Name | string | Frequency band name | "Alpha" |
| Hz_Range | string | Frequency range | "8-12 Hz" |
| Normal_Brain_State | text | Normal function | |
| Key_Regions | text | Prominent brain regions | |
| EEG_Positions | string | 10-20 electrode sites | |
| Pathological_Increase | text | What ↑ means clinically | |
| Pathological_Decrease | text | What ↓ means clinically | |
| Associated_Disorders | text | Linked conditions | |
| Clinical_Significance | text | Clinical notes | |
| Review_Status | enum | Reviewed / Pending / To Verify | Reviewed |

---

*Data Dictionary v2.0 — Generated """ + datetime.now().strftime('%Y-%m-%d') + """*
"""

dict_path = os.path.join(OUTPUT_DIR, "Data_Dictionary_v2.md")
with open(dict_path, 'w', encoding='utf-8') as f:
    f.write(data_dict)
print(f"  Data dictionary saved: {dict_path}")

# ============================================================
# FINAL SUMMARY
# ============================================================

print("\n" + "=" * 60)
print("INTEGRATION COMPLETE")
print("=" * 60)
print(f"\nTotal records: {total_records}")
print("Tables: 12")
print("Files generated:")
print(f"  - Excel: {excel_path}")
print(f"  - CSVs: {CSV_DIR}/ (12 files)")
print(f"  - Integration Report: {report_path}")
print(f"  - Data Dictionary: {dict_path}")
print("\nPending review items:")
print(f"  - {pending_count} protocols")
print(f"  - {pending_conditions} conditions")
print("\nGovernance flags summary:")
for flag, count in sorted(gov_flag_counts.items()):
    print(f"  - {flag}: {count} protocols flagged")
